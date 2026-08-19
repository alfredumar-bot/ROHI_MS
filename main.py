import os
import sys
import json
import sqlite3
import math
import random
import re
import logging
import calendar
import threading
import time
import shutil
from urllib.parse import urlparse
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from datetime import datetime, timedelta

# -------------------------------------------------------------
# AUTOMATIC LOG FILE GENERATOR
# -------------------------------------------------------------
APP_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_DIR = os.path.join(APP_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, "rohi_app.log")

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, mode='a', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger("ROHIApp")
logger.info("==========================================")
logger.info(f"ROHI Attendance App Starting at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
logger.info("==========================================")

# -------------------------------------------------------------
# Attendance report export protection
# -------------------------------------------------------------
# Password applied to exported attendance .xlsx reports (open-to-view lock).
REPORT_EXPORT_PASSWORD = "Rohi@3313@"

# -------------------------------------------------------------
# Attendance reminder schedule (Monday-Thursday only; Friday is WFH
# and Saturday/Sunday are non-working days - see
# ROHIApp._is_attendance_working_day).
# -------------------------------------------------------------
REMINDER_CHECKIN_HOUR, REMINDER_CHECKIN_MINUTE = 7, 50
REMINDER_CHECKOUT_HOUR, REMINDER_CHECKOUT_MINUTE = 15, 50
# Check-in nudges fire at exact clock times: 07:50 (first), 07:55 (second),
# 07:59 (third/last) - expressed as minutes after REMINDER_CHECKIN_MINUTE.
CHECKIN_REMINDER_OFFSETS_MINUTES = [0, 5, 9]
REMINDER_REPEAT_MINUTES = 5     # gap between check-out nudges
REMINDER_REPEAT_COUNT = 3       # extra check-out nudges fired after the on-time one
CHECKOUT_REMINDER_OFFSETS_MINUTES = [
    slot * REMINDER_REPEAT_MINUTES for slot in range(0, REMINDER_REPEAT_COUNT + 1)
]

# How often (seconds) the background auto-sync timer pushes pending rows to
# the PostgreSQL server without the user pressing "Synchronize Now".
AUTO_SYNC_INTERVAL_SECONDS = 180
# Report/file synchronization schedules. PostgreSQL/Google-Form sync can run
# frequently, while the Excel mirror jobs follow the requested ROHI schedule.
STAFF_EXCEL_SYNC_INTERVAL_SECONDS = 180  # all Excel auto-sync jobs every 3 minutes
ATTENDANCE_EXCEL_SYNC_INTERVAL_SECONDS = 180  # attendance Excel auto-sync every 3 minutes
TIMESHEET_EXCEL_SYNC_INTERVAL_SECONDS = 180  # timesheet Excel auto-sync every 3 minutes
LEAVE_EXCEL_SYNC_INTERVAL_SECONDS = 180  # leave Excel auto-sync every 3 minutes
EXCEL_SYNC_CONFIG_PATH = os.path.join(APP_DIR, "excel_sync_config.json")
EXCEL_SYNC_DEFAULTS = {
    "attendance_link": "https://docs.google.com/spreadsheets/d/e/2PACX-1vRt9XLMR2O9eW2ZMCNvhgxa2iSw8wZTIQZdh4mPNjj7D20YhiuAJSWgOTL3bpBm0g/pubhtml?gid=1377847618&single=true",
    "timesheet_link": "https://drive.google.com/drive/folders/1GTYacKygoa9O9vH_Oo--ZVZtCijKrEfD?usp=sharing",
    "leave_link": "https://drive.google.com/drive/folders/1H2EPqb3mPXB2Dty5o7bsg7gopO8cXOSH?usp=sharing",
    "staff_link": "https://docs.google.com/spreadsheets/d/e/2PACX-1vRaprW63u3MXCbO5RJ0v7xXKkmNp8rVt8JSpPtZupBUAHq38e41c6_laoLjyEfItA/pubhtml?gid=2005932240&single=true",
    # A Google Drive folder URL is a browse/share link, not an upload API.
    # Put a server/Apps Script upload endpoint in these optional fields to
    # enable real automatic file transfer.
    "attendance_endpoint": "https://script.google.com/macros/s/AKfycbwfqgtu86i3NqKUdR4oSX65urWJIKq_fNL6WqcJ4FEaz-P8bmb53CPZwkeou5_BYEqc/exec",
    "timesheet_endpoint": "https://script.google.com/macros/s/AKfycbwfqgtu86i3NqKUdR4oSX65urWJIKq_fNL6WqcJ4FEaz-P8bmb53CPZwkeou5_BYEqc/exec",
    "leave_endpoint": "https://script.google.com/macros/s/AKfycbwfqgtu86i3NqKUdR4oSX65urWJIKq_fNL6WqcJ4FEaz-P8bmb53CPZwkeou5_BYEqc/exec",
    "staff_endpoint": "https://script.google.com/macros/s/AKfycbwfqgtu86i3NqKUdR4oSX65urWJIKq_fNL6WqcJ4FEaz-P8bmb53CPZwkeou5_BYEqc/exec",
    "last_staff_sync": "",
    "last_attendance_sync": "",
    "last_timesheet_sync": "",
    "last_leave_sync": "",
    "link_status": {"attendance": False, "timesheet": False, "leave": False, "staff": False},
}

# -------------------------------------------------------------
# Public holidays used to label the timesheet grid as "PUBLIC HOLIDAY"
# instead of a normal work day. Add entries as "YYYY-MM-DD": "Name".
# Left empty by default - fill in ROHI's official holiday calendar for
# each year so the exported timesheet marks the right dates.
# -------------------------------------------------------------
ROHI_PUBLIC_HOLIDAYS = {
    # "2026-06-12": "Democracy Day",
}



def _load_excel_sync_config():
    data = dict(EXCEL_SYNC_DEFAULTS)
    try:
        if os.path.exists(EXCEL_SYNC_CONFIG_PATH):
            with open(EXCEL_SYNC_CONFIG_PATH, "r", encoding="utf-8") as f:
                saved = json.load(f)
            if isinstance(saved, dict):
                data.update(saved)
    except Exception:
        logger.exception("Could not load Excel sync configuration.")
    return data


def _save_excel_sync_config(data):
    try:
        with open(EXCEL_SYNC_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        return True
    except Exception:
        logger.exception("Could not save Excel sync configuration.")
        return False


def _http_upload_excel(filepath, endpoint, report_type):
    """Upload an XLSX file to a configured server/Apps-Script endpoint.

    The endpoint receives a multipart/form-data POST with fields:
    report_type, filename and file. Folder/share URLs are deliberately not
    treated as upload endpoints because Google Drive folder links do not accept
    anonymous file uploads.
    """
    if not endpoint:
        return False, "No upload endpoint configured."
    if not os.path.exists(filepath):
        return False, "Export file does not exist."
    try:
        filename = os.path.basename(filepath)
        # Google Apps Script Web Apps receive a JSON body reliably from Android.
        # Use this mode automatically for script.google.com endpoints; normal
        # HTTP upload endpoints continue to use multipart/form-data below.
        if "script.google.com" in endpoint.lower():
            import base64
            with open(filepath, "rb") as f:
                payload = base64.b64encode(f.read()).decode("ascii")
            body = json.dumps({
                "report_type": report_type,
                "filename": filename,
                "file_base64": payload,
            }).encode("utf-8")
            req = Request(endpoint, data=body, method="POST")
            req.add_header("Content-Type", "application/json")
            req.add_header("User-Agent", "ROHI-Attendance-App/1.8")
            with urlopen(req, timeout=60) as response:
                status = getattr(response, "status", 200)
                response.read(2048)
            if status < 200 or status >= 300:
                return False, f"Upload endpoint returned HTTP {status}."
            return True, "Uploaded to Google Drive successfully."

        boundary = "----ROHIExcelSyncBoundary" + str(int(time.time() * 1000))
        with open(filepath, "rb") as f:
            payload = f.read()

        def part(name, value):
            return (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
                f"{value}\r\n"
            ).encode("utf-8")

        body = bytearray()
        body.extend(part("report_type", report_type))
        body.extend(part("filename", filename))
        body.extend(
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
            f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
            .encode("utf-8")
        )
        body.extend(payload)
        body.extend(f"\r\n--{boundary}--\r\n".encode("utf-8"))

        req = Request(endpoint, data=bytes(body), method="POST")
        req.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
        req.add_header("User-Agent", "ROHI-Attendance-App/1.0")
        with urlopen(req, timeout=30) as response:
            status = getattr(response, "status", 200)
            response.read(1024)
        if status < 200 or status >= 300:
            return False, f"Upload endpoint returned HTTP {status}."
        return True, "Uploaded successfully."
    except HTTPError as exc:
        return False, f"Upload failed (HTTP {exc.code})."
    except URLError:
        return False, "Upload failed: no network connection."
    except Exception as exc:
        logger.exception("Excel upload failed:")
        return False, f"Upload failed: {exc}"


def _safe_filename(value):
    return re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "file")).strip("_") or "file"

def handle_unhandled_exception(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    logger.critical("Unhandled exception captured:", exc_info=(exc_type, exc_value, exc_traceback))

sys.excepthook = handle_unhandled_exception
# -------------------------------------------------------------

from kivy.clock import Clock, mainthread
from kivy.lang import Builder
from kivy.utils import platform
from kivy.metrics import dp
from kivy.core.window import Window

# Makes Kivy pan the screen so the focused text field stays visible above the
# on-screen keyboard, instead of the keyboard covering it (was hiding the
# password field during registration/login on most Android phones).
Window.softinput_mode = "below_target"

from kivymd.app import MDApp
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.label import MDLabel
from kivymd.uix.pickers import MDDatePicker
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton
from plyer import camera, gps

# Import RegistrationScreen from screens module
from registration_screen import RegistrationScreen
from reports_screen import ReportsScreen
from timesheet_screen import TimesheetScreen
from leave_screen import LeaveScreen
from settings_screen import SettingsScreen
from server_connection_screen import ServerConnectionScreen

import pg_sync
import gform_sync
import openpyxl

# Local database handlers
try:
    from database import (create_table, insert_staff, update_staff, get_staff_by_id, verify_login,
                           email_exists, staff_number_exists, create_leave_request,
                           get_leave_requests, get_leave_usage, get_leave_status_counts,
                           get_pending_gform_attendance, mark_gform_synced,
                           get_staff_count, clear_all_staff)
except ImportError:
    logger.warning("Local database module not found; creating inline fallback table handler.")
    
    def create_table():
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS staff (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fullname TEXT, sex TEXT, dob TEXT, blood_group TEXT, marital_status TEXT,
                nationality TEXT, state_origin TEXT, lga TEXT, address TEXT, next_of_kin TEXT,
                next_of_kin_phone TEXT, employment_type TEXT, state_office TEXT, cluster TEXT,
                department TEXT, section TEXT, position TEXT, staff_number TEXT, phone TEXT,
                email TEXT, facebook TEXT, twitter TEXT, instagram TEXT, telegram TEXT,
                linkedin TEXT, gps_coordinate TEXT, photo TEXT, password TEXT
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT,
                check_in_time TEXT,
                check_out_time TEXT,
                late_duration TEXT,
                attendance_status TEXT,
                gps_location TEXT,
                check_out_gps_location TEXT
            )
        ''')
        # Migration for databases created before check_out_gps_location existed.
        try:
            cursor.execute("ALTER TABLE attendance ADD COLUMN check_out_gps_location TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass  # Column already exists.
        conn.commit()
        conn.close()

    def get_leave_status_counts(staff_email):
        return {"Pending": 0, "Approved": 0, "Rejected": 0}

    def get_pending_gform_attendance(limit=50):
        return []

    def mark_gform_synced(attendance_id):
        pass

    def insert_staff(data):
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO staff (
                fullname, sex, dob, blood_group, marital_status, nationality, state_origin,
                lga, address, next_of_kin, next_of_kin_phone, employment_type, state_office,
                cluster, department, section, position, staff_number, phone, email,
                facebook, twitter, instagram, telegram, linkedin, gps_coordinate, photo, password
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            data.get("fullname"), data.get("sex"), data.get("dob"), data.get("blood_group"), data.get("marital_status"),
            data.get("nationality"), data.get("state_origin"), data.get("lga"), data.get("address"), data.get("next_of_kin"),
            data.get("next_of_kin_phone"), data.get("employment_type"), data.get("state_office"), data.get("cluster"),
            data.get("department"), data.get("section"), data.get("position"), data.get("staff_number"), data.get("phone"),
            data.get("email"), data.get("facebook"), data.get("twitter"), data.get("instagram"), data.get("telegram"),
            data.get("linkedin"), data.get("gps_coordinate"), data.get("photo"), data.get("password")
        ))
        conn.commit()
        conn.close()

    def verify_login(email_or_staff, password):
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        identifier = (email_or_staff or "").strip().lower()
        cursor.execute('''
            SELECT * FROM staff WHERE (LOWER(TRIM(email))=? OR LOWER(TRIM(staff_number))=?) AND password=?
        ''', (identifier, identifier, password))
        user = cursor.fetchone()
        conn.close()
        return user

    def get_staff_count():
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM staff")
        count = cursor.fetchone()[0]
        conn.close()
        return int(count or 0)

    def clear_all_staff():
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM staff")
        conn.commit()
        conn.close()


# Fixed office coordinates used for geofencing. Instead of relying on the GPS
# captured on a staff member's device at registration time (which could be
# inaccurate or spoofed), the registration form's GPS Coordinate field is
# auto-filled - and stays static - from whichever State Office is selected.
STATE_OFFICES = [
    "Borno State Office",
    "Adamawa HQ",
    "Yobe State Office",
    "Taraba State Office",
    "Benue State Office",
    "Sokoto State Office",
]

OFFICES = {
    "Borno State Office": {"latitude": 11.797352, "longitude": 13.143040, "radius": 100},
    "Adamawa HQ": {"latitude": 9.2781640, "longitude": 12.432640, "radius": 100},
    "Yobe State Office": {"latitude": 11.7460, "longitude": 11.9660, "radius": 100},
    "Taraba State Office": {"latitude": 8.8936, "longitude": 11.3595, "radius": 100},
    "Benue State Office": {"latitude": 7.7322, "longitude": 8.5391, "radius": 100},
    "Sokoto State Office": {"latitude": 13.0622, "longitude": 5.2339, "radius": 100},
}

# Compulsory registration fields (marked with * on the form), paired with
# the human-readable label shown in the "missing field" error dialog.
REQUIRED_REGISTRATION_FIELDS = [
    ("fullname", "Name"),
    ("sex", "Sex"),
    ("dob", "Date of Birth"),
    ("marital_status", "Marital Status"),
    ("nationality", "Nationality"),
    ("state_origin", "State of Origin"),
    ("lga", "LGA"),
    ("address", "Residential Address"),
    ("next_of_kin", "Next of Kin"),
    ("next_of_kin_phone", "Next of Kin Phone"),
    ("state_office", "State"),
    ("department", "Department"),
    ("section", "Section"),
    ("position", "Position"),
    ("employment_type", "Employment Type"),
    ("email", "Office Email"),
    ("phone", "Phone No"),
    ("staff_number", "Staff ID No"),
]


class SplashScreen(MDScreen):
    pass

class LoginScreen(MDScreen):
    pass

class DashboardScreen(MDScreen):
    pass


class ROHIAttendanceApp(MDApp):
    def build(self):
        self.title = "ROHI Staff App"
        self.theme_cls.primary_palette = "Green"
        
        logger.info("Initializing database table...")
        try:
            create_table()
            logger.info("Database initialized successfully.")
        except Exception:
            logger.exception("Error during database creation:")
        
        self.photo_path = ""
        self.signature_path = ""
        self.static_gps = "11.797352° N, 13.143040° E"
        # Never use the static office coordinate as a substitute for a live
        # phone GPS fix during Check-In/Check-Out.
        self.current_location = ""
        self._checkin_gps_fix = None
        self._checkout_gps_fix = None
        self._android_location_listener = None
        self._gps_poll_stop = None
        self._gps_timeout_event = None
        self.active_menu = None
        self.checked_in = False
        self.check_in_datetime = None
        self.check_out_datetime = None
        self.current_user = None
        self._session_restore_in_progress = False
        self._excel_sync_lock = threading.Lock()
        self._excel_sync_state = _load_excel_sync_config()
        self._last_attendance_excel_sync_date = ""
        self._last_staff_excel_sync_at = ""
        self._last_timesheet_excel_sync_at = ""
        self._last_leave_excel_sync_at = ""
        self._excel_leave_schedule_running = False
        self._excel_staff_schedule_running = False
        self._excel_attendance_schedule_running = False
        self._excel_timesheet_schedule_running = False
        self.late_duration_str = "On Time"
        self._checkin_pending = False
        self._checkout_pending = False
        # registered base GPS coordinate for Check-In to be accepted.
        # Editable from Settings > GPS/Geofencing; kept as an app-level attribute
        # (not hardcoded in _check_geofence) so a saved change takes effect immediately.
        # Persisted in app_settings.json (separate from server_config.json) so it
        # survives restarts but is untouched by Reset Database / server config changes.
        self.GEOFENCE_RADIUS_METERS = self._load_geofence_radius()

        # Load KV Files safely
        try:
            kv_files = ["splash.kv", "login.kv", "registration.kv", "dashboard.kv",
                        "reports.kv", "timesheet.kv", "leave.kv", "settings.kv", "server_connection.kv"]
            for kv_name in kv_files:
                kv_file = os.path.join(APP_DIR, kv_name)
                if os.path.exists(kv_file):
                    Builder.load_file(kv_file)
                else:
                    raise FileNotFoundError(f"Missing KV file: {kv_file}")
            logger.info("KV layout files loaded successfully.")
        except Exception:
            logger.exception("Failed to load KV layout files:")

        sm = MDScreenManager()
        sm.add_widget(SplashScreen(name="splash"))
        sm.add_widget(LoginScreen(name="login"))
        sm.add_widget(RegistrationScreen(name="registration"))
        sm.add_widget(DashboardScreen(name="dashboard"))
        sm.add_widget(ReportsScreen(name="reports"))
        sm.add_widget(TimesheetScreen(name="timesheet"))
        sm.add_widget(LeaveScreen(name="leave"))
        sm.add_widget(SettingsScreen(name="settings"))
        sm.add_widget(ServerConnectionScreen(name="server_connection"))

        self.registration_screen = sm.get_screen("registration")
        self.dashboard_screen = sm.get_screen("dashboard")
        self.reports_screen = sm.get_screen("reports")
        self.timesheet_screen = sm.get_screen("timesheet")
        self.leave_screen = sm.get_screen("leave")
        self.settings_screen = sm.get_screen("settings")
        self.server_connection_screen = sm.get_screen("server_connection")

        Clock.schedule_once(self._restore_or_open_login, 1.2)
        # Excel mirror jobs: staff corrections are checked once every 24 hours;
        # attendance is mirrored at 5:00 PM each day. The check also runs after
        # app launch so a phone that was offline at 5 PM catches up later.
        Clock.schedule_once(lambda dt: self._excel_sync_schedule_tick(), 3)
        Clock.schedule_interval(lambda dt: self._excel_sync_schedule_tick(), 30)

        # Run the live dashboard clock app-wide from launch (not tied to login),
        # so it's never stuck on the "Loading Date & Time..." placeholder.
        self.update_dashboard_time()
        Clock.schedule_interval(self.update_dashboard_time, 1)

        # Attendance Check-In/Check-Out reminders (07:50 / 15:50, Mon-Thu only)
        self._reminder_state = {"date": None, "checkin": set(), "checkout": set()}
        Clock.schedule_interval(self._reminder_tick, 30)

        # Silent background sync - pushes any pending rows to the server
        # every AUTO_SYNC_INTERVAL_SECONDS without needing "Synchronize Now".
        Clock.schedule_interval(self._auto_sync_tick, AUTO_SYNC_INTERVAL_SECONDS)

        return sm

    def on_start(self):
        """Dropdown menu fields are opened via on_focus handlers declared directly
        in kv/registration.kv. (Previously this also bound duplicate on_touch_down
        handlers here, which could double-trigger a menu open on a single tap and
        occasionally left the wrong field focused/populated - removed for that reason.)"""
        self._start_background_reminder_service()

    def _start_background_reminder_service(self):
        """Starts the ServiceReminder Android service (service_reminder.py) so
        the 07:50/07:55/07:59 (and 15:50-series) reminders still fire even
        while the app is closed or the screen is off - the in-app Clock-based
        _reminder_tick only runs while this app is actually in the foreground.
        No-op (silently) on desktop or if anything about the service API is
        unavailable, so it can never crash app startup."""
        if platform != "android":
            return
        try:
            from jnius import autoclass
            package_name = "org.rohi.rohiattendance"  # buildozer.spec: package.domain + package.name
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            ServiceReminder = autoclass("{}.ServiceReminder".format(package_name))
            ServiceReminder.start(PythonActivity.mActivity, "")
            self._reminder_service = ServiceReminder
            logger.info("Background reminder service started.")
        except Exception:
            logger.exception("Could not start background reminder service; falling back to in-app reminders only.")

    def _dismiss_active_menu(self):
        if self.active_menu:
            self.active_menu.dismiss()
            self.active_menu = None

    # -----------------------------
    # Local App Settings (app_settings.json) - geofence radius, etc.
    # -----------------------------
    def _app_settings_path(self):
        return os.path.join(os.path.dirname(__file__), "app_settings.json")

    def _load_reminders_enabled(self):
        try:
            path = self._app_settings_path()
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                return bool(data.get("reminders_enabled", True))
        except Exception:
            logger.exception("Failed to load app_settings.json; defaulting reminders to ON.")
        return True

    def _save_reminders_enabled(self, enabled):
        try:
            path = self._app_settings_path()
            data = {}
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            data["reminders_enabled"] = bool(enabled)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f)
        except Exception:
            logger.exception("Failed to save reminder notification preference:")

    def toggle_reminder_notifications(self, enabled):
        """Bound to the Attendance Reminder Notifications switch in Settings.
        Also read by service_reminder.py (the background service) so turning
        this off silences both the in-app and background reminders."""
        self._save_reminders_enabled(enabled)

    def open_notification_settings(self):
        """Bound to 'Open Notification Settings' in Settings - jumps straight
        to this app's Android system notification settings page, so the user
        can control sound/vibration/banner style at the OS level."""
        if platform != "android":
            return
        try:
            from jnius import autoclass
            Intent = autoclass("android.content.Intent")
            Settings = autoclass("android.provider.Settings")
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            activity = PythonActivity.mActivity
            intent = Intent(Settings.ACTION_APP_NOTIFICATION_SETTINGS)
            intent.putExtra(Settings.EXTRA_APP_PACKAGE, activity.getPackageName())
            activity.startActivity(intent)
        except Exception:
            logger.exception("Could not open Android notification settings:")

    def _load_geofence_radius(self):
        try:
            path = self._app_settings_path()
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                return int(data.get("geofence_radius_meters", 100))
        except Exception:
            logger.exception("Failed to load app_settings.json; using default geofence radius.")
        return 100

    def _load_google_form_url(self):
        try:
            path = self._app_settings_path()
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                return str(data.get("google_form_url", ""))
        except Exception:
            logger.exception("Failed to load app_settings.json; using empty Google Form URL.")
        return ""

    def _save_google_form_url(self, url):
        try:
            path = self._app_settings_path()
            data = {}
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            data["google_form_url"] = url
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f)
        except Exception:
            logger.exception("Failed to save Google Form URL to app_settings.json:")

    def open_google_form(self):
        """Opens the staff-configured Google Form link in the phone's browser."""
        ids = self.server_connection_screen.ids
        url = ids.google_form_url_field.text.strip() if hasattr(ids, 'google_form_url_field') else ""
        if not url:
            self._set_sync_result_label("Enter a Google Form URL first.", ok=False)
            return
        if not (url.startswith("http://") or url.startswith("https://")):
            url = "https://" + url
        self._save_google_form_url(url)
        if platform != 'android':
            self._set_sync_result_label("Opening a browser is only available on an Android device.", ok=False)
            return
        try:
            from jnius import autoclass
            Intent = autoclass('android.content.Intent')
            Uri = autoclass('android.net.Uri')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            intent = Intent(Intent.ACTION_VIEW, Uri.parse(url))
            PythonActivity.mActivity.startActivity(intent)
        except Exception as e:
            logger.exception("Could not open Google Form URL:")
            self._set_sync_result_label(f"Could not open the link: {e}", ok=False)

    def save_google_form_prefilled_url(self):
        """Reads the pre-filled Google Form link pasted into Settings, parses
        out its entry IDs by matching the ROHI_* placeholder tokens, and saves
        the result so the app can silently POST attendance rows straight into
        the form (see gform_sync.py for the full field-token scheme)."""
        ids = self.server_connection_screen.ids
        url = (ids.google_form_prefilled_url_field.text.strip()
               if hasattr(ids, 'google_form_prefilled_url_field') else "")

        ok, message, entry_map, response_url = gform_sync.parse_prefilled_url(url)
        status_label = getattr(ids, 'google_form_auto_sync_status_label', None)

        if not ok:
            if status_label is not None:
                status_label.text = message
                status_label.text_color = (0.8, 0.1, 0.1, 1)
            logger.warning("Google Form auto-detect failed: %s", message)
            return

        config = {
            "response_url": response_url,
            "entry_map": entry_map,
            "configured_fields": list(entry_map.keys()),
        }
        saved, save_message = gform_sync.save_config(config)
        if status_label is not None:
            status_label.text = message if saved else save_message
            status_label.text_color = (0.1, 0.5, 0.15, 1) if saved else (0.8, 0.1, 0.1, 1)
        logger.info("Google Form auto-sync configured: %s", message)

    def _load_google_form_auto_sync_status(self):
        """Restores the auto-sync status label when the Server Connection
        screen is (re)opened, so the admin can see it's still configured
        without having to paste the link again."""
        try:
            ids = self.server_connection_screen.ids
            status_label = getattr(ids, 'google_form_auto_sync_status_label', None)
            if status_label is None:
                return
            if gform_sync.is_configured():
                config = gform_sync.load_config()
                fields = ", ".join(config.get("configured_fields", []))
                status_label.text = f"Auto-submit is ON ({fields})."
                status_label.text_color = (0.1, 0.5, 0.15, 1)
            else:
                status_label.text = "Auto-submit is OFF - paste a pre-filled link above."
                status_label.text_color = (0.5, 0.5, 0.5, 1)
        except Exception:
            logger.exception("Failed to load Google Form auto-sync status:")

    # -----------------------------
    # Attendance History Reset
    # -----------------------------
    def reset_database_tables(self):
        """Clear attendance history only; NEVER delete staff registrations."""
        db_path = os.path.join(APP_DIR, "attendance.db")
        conn = sqlite3.connect(db_path, timeout=15)
        try:
            conn.execute("PRAGMA busy_timeout=15000")
            try: conn.execute("PRAGMA journal_mode=WAL")
            except sqlite3.DatabaseError: pass
            conn.execute("DELETE FROM attendance")
            conn.commit()
            logger.info("All local attendance records/reports were cleared; staff records preserved.")
        finally: conn.close()
        create_table()

    def confirm_reset_database(self):
        """Confirm clearing attendance history while preserving all staff."""
        self._dismiss_active_menu()
        if getattr(self, "_reset_dialog", None): self._reset_dialog.dismiss()
        self._reset_dialog = MDDialog(
            title="Clear Attendance History?",
            text="This permanently deletes local check-in/check-out records and attendance reports. Registered staff, profiles, passwords and office GPS settings will NOT be deleted. Continue?",
            buttons=[
                MDFlatButton(text="CANCEL", on_release=lambda *a: self._reset_dialog.dismiss()),
                MDFlatButton(text="CLEAR ATTENDANCE", theme_text_color="Custom", text_color=(0.8,0.1,0.1,1), on_release=lambda *a: self._do_confirmed_reset()),
            ],)
        self._reset_dialog.open()

    def _do_confirmed_reset(self):
        if getattr(self, "_reset_dialog", None): self._reset_dialog.dismiss()
        self.reset_database_tables()
        try: self.update_dashboard_metrics()
        except Exception: logger.exception("Could not refresh attendance summary after clearing history.")
        try:
            if hasattr(self, "settings_screen") and "reset_status_label" in self.settings_screen.ids:
                self.settings_screen.ids.reset_status_label.text = "Attendance history cleared. Staff registrations were preserved."
        except Exception: pass

    # -----------------------------
    # Settings / Server Connection navigation
    # -----------------------------
    def open_settings(self):
        self._dismiss_active_menu()
        if hasattr(self, "settings_screen") and hasattr(self.settings_screen, "ids"):
            if "server_status_label" in self.settings_screen.ids:
                status = pg_sync.get_status()
                state = "Connected" if status["connected"] else "Not connected"
                lbl = self.settings_screen.ids.server_status_label
                lbl.text = f"{state} · {status['pending_count']} record(s) pending sync"
                lbl.theme_text_color = "Custom"
                lbl.text_color = (0.1, 0.6, 0.2, 1) if status["connected"] else (0.6, 0.15, 0.15, 1)
            if "reminders_switch" in self.settings_screen.ids:
                self.settings_screen.ids.reminders_switch.active = self._load_reminders_enabled()
        self.root.current = "settings"

    def open_server_connection(self):
        self._dismiss_active_menu()
        self.refresh_server_connection_screen()
        self.root.current = "server_connection"
        # Test all configured report links in parallel so their individual
        # green/red status badges reflect the real connection state.
        Clock.schedule_once(lambda dt: self._auto_test_excel_links(), 0.2)

    def _auto_test_excel_links(self):
        for report_type in ("attendance", "timesheet", "leave", "staff"):
            self.connect_excel_report(report_type)

    def refresh_server_connection_screen(self):
        """Populates the Server Connection screen fields/status from saved config + live state."""
        if not (hasattr(self, "server_connection_screen") and hasattr(self.server_connection_screen, "ids")):
            return
        ids = self.server_connection_screen.ids
        config = pg_sync.load_config()

        field_map = {
            "server_name_field": "server_name",
            "host_field": "host",
            "port_field": "port",
            "dbname_field": "dbname",
            "username_field": "username",
            "password_field": "password",
        }
        for widget_id, key in field_map.items():
            if widget_id in ids:
                ids[widget_id].text = str(config.get(key) or "")

        if "sslmode_field" in ids:
            ids.sslmode_field.text = config.get("sslmode") or "prefer"

        if "google_form_url_field" in ids:
            ids.google_form_url_field.text = self._load_google_form_url()
        self._load_google_form_auto_sync_status()
        self._refresh_excel_sync_fields()

        status = pg_sync.get_status()
        status_color = (0.1, 0.6, 0.2, 1) if status["connected"] else (0.6, 0.15, 0.15, 1)
        status_bg = (0.85, 0.93, 0.85, 1) if status["connected"] else (0.93, 0.85, 0.85, 1)
        if "connection_status_label" in ids:
            ids.connection_status_label.text = "Connected" if status["connected"] else "Not Connected"
            ids.connection_status_label.theme_text_color = "Custom"
            ids.connection_status_label.text_color = status_color
        if "connection_status_icon" in ids:
            ids.connection_status_icon.icon = "check-circle" if status["connected"] else "close-circle"
            ids.connection_status_icon.text_color = status_color
        if "connection_status_badge" in ids:
            ids.connection_status_badge.md_bg_color = status_bg
        if "last_sync_label" in ids:
            ids.last_sync_label.text = f"Last Sync: {status['last_sync_time']}"
        if "sync_result_label" in ids:
            ids.sync_result_label.text = f"{status['pending_count']} record(s) pending sync."

    def _collect_server_connection_form(self):
        ids = self.server_connection_screen.ids
        port_text = ids.port_field.text.strip() if "port_field" in ids else "5432"
        try:
            port = int(port_text) if port_text else 5432
        except ValueError:
            port = 5432
        return {
            "server_name": ids.server_name_field.text.strip() if "server_name_field" in ids else "",
            "host": ids.host_field.text.strip() if "host_field" in ids else "",
            "port": port,
            "dbname": ids.dbname_field.text.strip() if "dbname_field" in ids else "",
            "username": ids.username_field.text.strip() if "username_field" in ids else "",
            "password": ids.password_field.text if "password_field" in ids else "",
            "sslmode": ids.sslmode_field.text.strip() if "sslmode_field" in ids else "prefer",
        }

    # -----------------------------
    # Excel link / folder connection
    # -----------------------------
    def _set_excel_link_status(self, report_type, connected, message=""):
        """Update one report link's green/red connection indicator."""
        try:
            ids = self.server_connection_screen.ids
            icon_id = f"{report_type}_connection_icon"
            label_id = f"{report_type}_connection_label"
            badge_id = f"{report_type}_connection_badge"
            color = (0.10, 0.60, 0.20, 1) if connected else (0.80, 0.10, 0.10, 1)
            bg = (0.86, 0.95, 0.87, 1) if connected else (0.96, 0.87, 0.87, 1)
            if icon_id in ids:
                ids[icon_id].icon = "check-circle" if connected else "close-circle"
                ids[icon_id].text_color = color
            if label_id in ids:
                ids[label_id].text = "Connected" if connected else "Not Connected"
                ids[label_id].theme_text_color = "Custom"
                ids[label_id].text_color = color
            if badge_id in ids:
                ids[badge_id].md_bg_color = bg
            self._excel_sync_state.setdefault("link_status", {})[report_type] = bool(connected)
            if message and hasattr(ids, "excel_sync_result_label"):
                ids.excel_sync_result_label.text = message
            _save_excel_sync_config(self._excel_sync_state)
            self._update_excel_sync_dashboard_status()
        except Exception:
            logger.exception("Could not update Excel link status for %s", report_type)

    def connect_excel_report(self, report_type):
        """Test one report link in the background and show green/red status."""
        url = str(self._excel_sync_state.get(f"{report_type}_link") or "").strip()
        if not url.startswith(("http://", "https://")):
            self._set_excel_link_status(report_type, False, f"{report_type.title()} link is missing or invalid.")
            return
        self._set_excel_link_status(report_type, False, f"Connecting to {report_type.title()} link...")
        def worker():
            try:
                req = Request(url, headers={"User-Agent": "ROHI-Attendance-App/1.7"}, method="GET")
                with urlopen(req, timeout=10) as response:
                    status = getattr(response, "status", 200)
                    response.read(256)
                ok = 200 <= status < 400
                msg = f"{report_type.title()} link connected." if ok else f"{report_type.title()} link failed (HTTP {status})."
            except Exception as exc:
                logger.exception("Excel link connection test failed for %s", report_type)
                ok = False
                msg = f"{report_type.title()} link failed: {exc}"
            Clock.schedule_once(lambda dt: self._set_excel_link_status(report_type, ok, msg), 0)
        threading.Thread(target=worker, daemon=True).start()

    def _refresh_excel_sync_fields(self):
        if not hasattr(self, "server_connection_screen"):
            return
        ids = self.server_connection_screen.ids
        mapping = {
            "attendance_link_field": "attendance_link",
            "timesheet_link_field": "timesheet_link",
            "leave_link_field": "leave_link",
            "staff_link_field": "staff_link",
            "attendance_endpoint_field": "attendance_endpoint",
            "timesheet_endpoint_field": "timesheet_endpoint",
            "leave_endpoint_field": "leave_endpoint",
            "staff_endpoint_field": "staff_endpoint",
        }
        for widget_id, key in mapping.items():
            if widget_id in ids:
                ids[widget_id].text = str(self._excel_sync_state.get(key) or "")
        statuses = self._excel_sync_state.get("link_status", {})
        for report_type in ("attendance", "timesheet", "leave", "staff"):
            self._set_excel_link_status(report_type, bool(statuses.get(report_type, False)))
        self._update_excel_sync_dashboard_status()

    def _collect_excel_sync_form(self):
        ids = self.server_connection_screen.ids
        data = dict(self._excel_sync_state)
        mapping = {
            "attendance_link_field": "attendance_link",
            "timesheet_link_field": "timesheet_link",
            "leave_link_field": "leave_link",
            "staff_link_field": "staff_link",
            "attendance_endpoint_field": "attendance_endpoint",
            "timesheet_endpoint_field": "timesheet_endpoint",
            "leave_endpoint_field": "leave_endpoint",
            "staff_endpoint_field": "staff_endpoint",
        }
        for widget_id, key in mapping.items():
            if widget_id in ids:
                data[key] = ids[widget_id].text.strip()
        return data

    def save_excel_sync_settings(self):
        data = self._collect_excel_sync_form()
        self._excel_sync_state = data
        ok = _save_excel_sync_config(data)
        message = "Excel links saved." if ok else "Could not save Excel links."
        if hasattr(self.server_connection_screen.ids, "excel_sync_result_label"):
            self.server_connection_screen.ids.excel_sync_result_label.text = message
        self._update_excel_sync_dashboard_status()
        return ok

    def connect_excel_sync(self):
        data = self._collect_excel_sync_form()
        links = [data.get(k, "") for k in (
            "attendance_link", "timesheet_link", "leave_link", "staff_link"
        )]
        valid = all(u.startswith(("http://", "https://")) for u in links if u)
        if not valid:
            if hasattr(self.server_connection_screen.ids, "excel_sync_result_label"):
                self.server_connection_screen.ids.excel_sync_result_label.text = "Use valid http:// or https:// links."
            return False
        self._excel_sync_state = data
        _save_excel_sync_config(data)
        endpoint_count = sum(bool(data.get(k)) for k in (
            "attendance_endpoint", "timesheet_endpoint", "leave_endpoint", "staff_endpoint"
        ))
        message = (
            f"Connected: {len([u for u in links if u])}/4 report links saved. "
            f"{endpoint_count}/4 upload endpoints configured."
        )
        if hasattr(self.server_connection_screen.ids, "excel_sync_result_label"):
            self.server_connection_screen.ids.excel_sync_result_label.text = message
        self._update_excel_sync_dashboard_status()
        return True

    def open_excel_sync_link(self, report_type):
        url = str(self._excel_sync_state.get(f"{report_type}_link") or "").strip()
        if not url:
            return
        if platform != "android":
            return
        try:
            from jnius import autoclass
            Intent = autoclass("android.content.Intent")
            Uri = autoclass("android.net.Uri")
            activity = self._get_android_activity()
            activity.startActivity(Intent(Intent.ACTION_VIEW, Uri.parse(url)))
        except Exception as exc:
            logger.exception("Could not open Excel sync link:")
            if hasattr(self.server_connection_screen.ids, "excel_sync_result_label"):
                self.server_connection_screen.ids.excel_sync_result_label.text = f"Could not open link: {exc}"

    def _update_excel_sync_dashboard_status(self):
        try:
            if not hasattr(self, "dashboard_screen"):
                return
            ids = self.dashboard_screen.ids
            statuses = self._excel_sync_state.get("link_status", {})
            connected_count = sum(bool(statuses.get(k, False)) for k in ("attendance", "timesheet", "leave", "staff"))
            if "excel_sync_status" in ids:
                if connected_count == 4:
                    ids.excel_sync_status.text = "Excel Sync: CONNECTED (4/4)"
                    ids.excel_sync_status.text_color = (0.13, 0.40, 0.16, 1)
                elif connected_count > 0:
                    ids.excel_sync_status.text = f"Excel Sync: {connected_count}/4 CONNECTED"
                    ids.excel_sync_status.text_color = (0.75, 0.55, 0.05, 1)
                else:
                    ids.excel_sync_status.text = "Excel Sync: NOT CONNECTED"
                    ids.excel_sync_status.text_color = (0.8, 0.1, 0.1, 1)
        except Exception:
            logger.exception("Could not update Excel sync dashboard status.")

    # -----------------------------
    # Server Connection actions
    # -----------------------------
    # NOTE: pg8000 network calls (test/connect/sync) are blocking. Running them
    # directly on the on_release handler used to freeze the whole app - no
    # button feedback, no repaint - for as long as the TCP attempt took (which
    # can be 30s+ against an unreachable host), which looked exactly like "the
    # server connection isn't working" even when the credentials were fine.
    # Every action below now runs on a background thread; only the quick,
    # local UI updates (busy state, then the result) touch widgets, and those
    # are marshalled back to the main thread with @mainthread.
    def _set_server_connection_busy(self, busy, busy_text=None):
        if not (hasattr(self, "server_connection_screen") and hasattr(self.server_connection_screen, "ids")):
            return
        ids = self.server_connection_screen.ids
        for btn_id in ("test_connection_btn", "save_settings_btn", "connect_btn",
                       "disconnect_btn", "sync_now_btn"):
            if btn_id in ids:
                ids[btn_id].disabled = busy
                ids[btn_id].opacity = 0.5 if busy else 1
        if busy and busy_text and "connection_status_label" in ids:
            ids.connection_status_label.text = busy_text
            ids.connection_status_label.theme_text_color = "Custom"
            ids.connection_status_label.text_color = (0.75, 0.55, 0.05, 1)  # amber = in progress
            if "connection_status_icon" in ids:
                ids.connection_status_icon.icon = "timer-sand"
                ids.connection_status_icon.text_color = (0.75, 0.55, 0.05, 1)
            if "connection_status_badge" in ids:
                ids.connection_status_badge.md_bg_color = (0.96, 0.92, 0.82, 1)

    def _run_server_action(self, busy_text, worker_fn, on_done_fn):
        """Runs worker_fn() on a background thread, then marshals on_done_fn
        (its return value) back to the UI thread."""
        self._set_server_connection_busy(True, busy_text)

        def _worker():
            try:
                result = worker_fn()
            except Exception as e:
                logger.exception("Server connection action failed:")
                result = (False, f"Unexpected error: {e}")
            self._server_action_done(result, on_done_fn)

        threading.Thread(target=_worker, daemon=True).start()

    @mainthread
    def _server_action_done(self, result, on_done_fn):
        self._set_server_connection_busy(False)
        on_done_fn(result)

    def server_connection_test(self):
        config = self._collect_server_connection_form()
        self._run_server_action(
            "Testing connection...",
            lambda: pg_sync.test_connection(config),
            lambda result: self._on_test_connection_done(result),
        )

    def _on_test_connection_done(self, result):
        ok, message = result
        self._set_sync_result_label(message, ok=ok)
        # Testing doesn't change the live connection - restore the label to
        # reflect actual connection state rather than leaving "Testing..." up.
        self.refresh_server_connection_screen()

    def server_connection_save(self):
        config = self._collect_server_connection_form()
        ok, message = pg_sync.save_config(config)
        self._set_sync_result_label(message, ok=ok)

    def server_connection_connect(self):
        """Connect. Per the app's sync policy, a successful connect immediately
        triggers a sync of any pending local rows (auto-sync-on-reconnect)."""
        config = self._collect_server_connection_form()
        pg_sync.save_config(config)
        self._run_server_action(
            "Connecting...",
            lambda: pg_sync.connect(config, auto_sync=True),
            lambda result: self._on_connect_done(result),
        )

    def _on_connect_done(self, result):
        ok, message = result
        self._set_sync_result_label(message, ok=ok)
        self.refresh_server_connection_screen()

    def server_connection_disconnect(self):
        self._run_server_action(
            "Disconnecting...",
            lambda: pg_sync.disconnect(),
            lambda result: self._on_disconnect_done(result),
        )

    def _on_disconnect_done(self, result):
        ok, message = result
        self._set_sync_result_label(message, ok=ok)
        self.refresh_server_connection_screen()

    def server_connection_sync_now(self):
        config = self._collect_server_connection_form()
        self._run_server_action(
            "Synchronizing...",
            lambda: pg_sync.synchronize_now(config),
            lambda result: self._on_sync_now_done(result),
        )

    def _on_sync_now_done(self, result):
        ok, message = result[0], result[1]
        self._set_sync_result_label(message, ok=ok)
        self.refresh_server_connection_screen()

    def _set_sync_result_label(self, message, ok=None):
        if hasattr(self, "server_connection_screen") and hasattr(self.server_connection_screen, "ids"):
            ids = self.server_connection_screen.ids
            if "sync_result_label" in ids:
                ids.sync_result_label.text = message
                if ok is True:
                    ids.sync_result_label.theme_text_color = "Custom"
                    ids.sync_result_label.text_color = (0.13, 0.40, 0.16, 1)
                elif ok is False:
                    ids.sync_result_label.theme_text_color = "Custom"
                    ids.sync_result_label.text_color = (0.8, 0.1, 0.1, 1)
                else:
                    ids.sync_result_label.theme_text_color = "Secondary"
        logger.info(f"[Server Connection] {message}")

    # -----------------------------
    # Session persistence: remain logged in until the user explicitly logs out
    # -----------------------------
    def _session_file(self):
        return os.path.join(APP_DIR, "session.json")

    def _save_login_session(self, email):
        try:
            with open(self._session_file(), "w", encoding="utf-8") as f:
                json.dump({"email": str(email).strip().lower(), "saved_at": datetime.now().isoformat()}, f)
        except Exception:
            logger.exception("Could not save login session.")

    def _clear_login_session(self):
        try:
            path = self._session_file()
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            logger.exception("Could not clear login session.")

    def _restore_or_open_login(self, dt=None):
        if self.current_user:
            return
        try:
            path = self._session_file()
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                email = str(data.get("email") or "").strip().lower()
                if email:
                    db_path = os.path.join(APP_DIR, "attendance.db")
                    conn = sqlite3.connect(db_path, timeout=10)
                    try:
                        user = conn.execute(
                            "SELECT * FROM staff WHERE LOWER(TRIM(email)) = ? LIMIT 1", (email,)
                        ).fetchone()
                    finally:
                        conn.close()
                    if user:
                        self.current_user = self._ensure_unique_id(user)
                        self._save_login_session(email)
                        self._populate_dashboard_from_current_user()
                        self.verify_existing_checkin(email)
                        self.update_dashboard_metrics()
                        self.root.current = "dashboard"
                        logger.info("Restored authenticated session for %s.", email)
                        return
        except Exception:
            logger.exception("Could not restore login session; showing login screen.")
        self.root.current = "login"

    # -----------------------------
    # Navigation & Profile Edit Binding
    # -----------------------------
    def open_login(self, dt=None):
        self.root.current = "login"

    def open_registration(self):
        # A phone may hold only one staff registration. Editing the
        # already-logged-in profile is always allowed; starting a brand
        # new registration is blocked once any staff record exists locally.
        if not self.current_user:
            try:
                already_registered = get_staff_count() > 0
            except Exception:
                logger.exception("Failed to check existing staff count before registration.")
                already_registered = False

            if already_registered:
                self._show_registration_error(
                    "Already Registered",
                    "This phone already has a registered staff account.\n\n"
                    "Only one staff registration is allowed per phone. "
                    "Please log in with the existing account below, or "
                    "contact your administrator if this phone needs to be reset."
                )
                return

        self.populate_registration_for_edit()
        self.root.current = "registration"

    def populate_registration_for_edit(self):
        """Pre-fills registration screen fields with logged-in staff profile data."""
        if not self.current_user:
            return
        
        try:
            reg_ids = self.registration_screen.ids
            user = self.current_user

            field_mapping = {
                'fullname': user[1],
                'sex': user[2],
                'dob': user[3],
                'blood_group': user[4],
                'marital_status': user[5],
                'nationality': user[6],
                'state_origin': user[7],
                'lga': user[8],
                'address': user[9],
                'next_of_kin': user[10],
                'next_of_kin_phone': user[11],
                'employment_type': user[12],
                'state_office': user[13],
                'cluster': user[14],
                'department': user[15],
                'section': user[16],
                'position': user[17],
                'staff_number': user[18],
                'phone': user[19],
                'email': user[20],
                'facebook': user[21],
                'twitter': user[22],
                'instagram': user[23],
                'telegram': user[24],
                'linkedin': user[25],
                'gps_coordinate': user[26]
            }

            for field_id, value in field_mapping.items():
                if hasattr(reg_ids, field_id) and value is not None:
                    getattr(reg_ids, field_id).text = str(value)

            # Cache the office GPS coordinate on the app (used by submit_staff)
            # and show it in the read-only field so editing a profile doesn't
            # wipe the geofence coordinate tied to the staff member's office.
            self._selected_office_gps_coordinate = user[26] if len(user) > 26 and user[26] else ""
            if hasattr(reg_ids, 'gps_coordinate') and self._selected_office_gps_coordinate:
                reg_ids.gps_coordinate.text = self._selected_office_gps_coordinate

            # Genotype and Re-integrated fields were removed from the
            # registration form, so their stored values (if any, from
            # older records) are intentionally no longer displayed.

            if hasattr(reg_ids, 'photo_preview') and len(user) > 27 and user[27]:
                if os.path.exists(user[27]):
                    reg_ids.photo_preview.source = user[27]

            logger.info("Registration form pre-filled with active user profile.")
        except Exception:
            logger.exception("Error populating registration form for profile edit:")

    def open_dashboard(self):
        self.root.current = "dashboard"

    def logout(self):
        logger.info("User logged out.")
        self.checked_in = False
        self.check_in_datetime = None
        self.check_out_datetime = None
        self.current_user = None
        self._clear_login_session()
        self.root.current = "login"

    def forgot_password(self):
        logger.info("Forgot password triggered.")

    def _populate_dashboard_from_current_user(self):
        """Fills every dashboard profile field from self.current_user. Used both
        right after login and right after a profile edit is saved, so the two
        can never drift out of sync (this was the cause of edited profiles still
        showing old staff ID/email after saving)."""
        if not self.current_user:
            return
        user = self.current_user
        fullname = str(user[1]) if len(user) > 1 and user[1] else "Staff Member"
        position = str(user[17]) if len(user) > 17 and user[17] else "MEAL Officer"
        staff_num = str(user[18]) if len(user) > 18 and user[18] else "ROHI/MIU/P/067"
        unique_id = str(user[29]) if len(user) > 29 and user[29] else "-"
        email_val = str(user[20]) if len(user) > 20 and user[20] else ""
        state_office = str(user[13]) if len(user) > 13 and user[13] else ""
        cluster = str(user[14]) if len(user) > 14 and user[14] else ""
        registered_gps = str(user[26]) if len(user) > 26 and user[26] else self.static_gps
        photo_path = str(user[27]) if len(user) > 27 and user[27] else ""

        dash_ids = self.dashboard_screen.ids
        try:
            leave_counts = get_leave_status_counts(email_val)
            if hasattr(dash_ids, 'leave_pending_count'):
                dash_ids.leave_pending_count.text = str(leave_counts.get("Pending", 0))
            if hasattr(dash_ids, 'leave_approved_count'):
                dash_ids.leave_approved_count.text = str(leave_counts.get("Approved", 0))
            if hasattr(dash_ids, 'leave_rejected_count'):
                dash_ids.leave_rejected_count.text = str(leave_counts.get("Rejected", 0))
        except Exception:
            logger.exception("Unable to populate leave status counts on dashboard")
        if hasattr(dash_ids, 'welcome_label'):
            dash_ids.welcome_label.text = f"Welcome, {fullname}"
        if hasattr(dash_ids, 'role_label'):
            dash_ids.role_label.text = f"Position: {position}"
        if hasattr(dash_ids, 'email_label'):
            dash_ids.email_label.text = f"Email: {email_val}"
        if hasattr(dash_ids, 'staff_id_label'):
            dash_ids.staff_id_label.text = f"Staff ID: {staff_num}"
        if hasattr(dash_ids, 'unique_id_label'):
            dash_ids.unique_id_label.text = f"Unique ID: {unique_id}"
        if hasattr(dash_ids, 'state_office_label'):
            dash_ids.state_office_label.text = f"State Office: {state_office}" if state_office else "State Office: -"
        if hasattr(dash_ids, 'cluster_label'):
            dash_ids.cluster_label.text = f"Cluster: {cluster}" if cluster else "Cluster: -"
        if hasattr(dash_ids, 'static_gps_label'):
            dash_ids.static_gps_label.text = registered_gps
        if hasattr(dash_ids, 'current_gps_label'):
            dash_ids.current_gps_label.text = self.current_location
        if hasattr(dash_ids, 'dash_photo') and photo_path and os.path.exists(photo_path):
            dash_ids.dash_photo.source = photo_path
        self._update_excel_sync_dashboard_status()

        return email_val

    def perform_login(self):
        """Authenticates user and populates Dashboard UI with user records."""
        logger.info("perform_login triggered")
        try:
            login_screen = self.root.get_screen("login")
            email_or_staff = login_screen.ids.email.text.strip()
            password = login_screen.ids.password.text.strip()

            if not email_or_staff or not password:
                logger.warning("Login attempt with empty fields.")
                self._show_login_error("Please enter your email/staff number and password.")
                return

            logger.info(f"Authenticating login for user/staff ID: {email_or_staff}")
            user = verify_login(email_or_staff, password)

            if not user:
                logger.warning(f"Failed login attempt for '{email_or_staff}': Invalid credentials.")
                self._show_login_error("Incorrect email/staff number or password.")
                return

            logger.info(f"User '{email_or_staff}' authenticated successfully.")
            self.current_user = self._ensure_unique_id(user)
            self._save_login_session(str(self.current_user[20]) if len(self.current_user) > 20 else email_or_staff)

            # From here on, authentication has already succeeded - the user IS
            # logged in. Everything below is populating the dashboard, which
            # must NOT be allowed to strand the user back on the login screen
            # on failure (previously a single exception in this block - e.g. a
            # stale/partial local schema - would be silently swallowed by the
            # outer except and leave root.current on "login" with no feedback,
            # even though current_user was already set. That looked like "login
            # doesn't work" and could only be worked around by going through
            # Register > Edit Profile, which has its own success path to the
            # dashboard). So each step here is isolated and non-fatal.
            try:
                email_val = self._populate_dashboard_from_current_user() or email_or_staff
            except Exception:
                logger.exception("Failed to populate dashboard after login (non-fatal):")
                email_val = email_or_staff

            try:
                self.verify_existing_checkin(email_val)
            except Exception:
                logger.exception("Failed to verify existing check-in after login (non-fatal):")

            try:
                self.update_dashboard_metrics()
            except Exception:
                logger.exception("Failed to update dashboard metrics after login (non-fatal):")

            self.root.current = "dashboard"
        except Exception:
            logger.exception("Exception during perform_login:")
            self._show_login_error("Something went wrong logging in. Please try again.")

    def _show_login_error(self, message):
        """Surfaces login problems on-screen instead of failing silently -
        previously a failed/errored login attempt gave no visible feedback at
        all, so it just looked like the LOGIN button did nothing."""
        try:
            login_screen = self.root.get_screen("login")
            if hasattr(login_screen.ids, "login_error_label"):
                login_screen.ids.login_error_label.text = message
        except Exception:
            logger.exception("Failed to display login error label:")

    # -----------------------------
    # Geofencing & Daily Reset Helpers
    # -----------------------------
    @staticmethod
    def _parse_coordinate(coord_str):
        """Parses a 'lat° N, lon° E' style string into (lat, lon) floats.
        Handles S/W negatives too. Returns None if the string can't be parsed."""
        if not coord_str:
            return None
        try:
            import re
            matches = re.findall(r"(-?\d+\.?\d*)\s*°?\s*([NSEW]?)", coord_str.upper())
            nums = [m for m in matches if m[0]]
            if len(nums) < 2:
                return None
            lat = float(nums[0][0])
            if nums[0][1] == 'S':
                lat = -abs(lat)
            lon = float(nums[1][0])
            if nums[1][1] == 'W':
                lon = -abs(lon)
            return lat, lon
        except Exception:
            return None

    def _distance_meters(self, coord_a, coord_b):
        """Great-circle (Haversine) distance in meters between two coordinate strings.
        Returns None if either coordinate can't be parsed."""
        a = self._parse_coordinate(coord_a)
        b = self._parse_coordinate(coord_b)
        if not a or not b:
            return None

        lat1, lon1 = a
        lat2, lon2 = b
        R = 6371000  # Earth radius in meters

        phi1, phi2 = math.radians(lat1), math.radians(lat2)
        d_phi = math.radians(lat2 - lat1)
        d_lambda = math.radians(lon2 - lon1)

        h = (math.sin(d_phi / 2) ** 2
             + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2)
        return 2 * R * math.asin(min(1, math.sqrt(h)))

    def _get_android_activity(self):
        """Return the active Android Activity in both APK and Pydroid/Kivy runs."""
        candidates = (
            "org.kivy.android.PythonActivity",
            "org.renpy.android.PythonActivity",
        )
        last_error = None
        try:
            from jnius import autoclass
            for name in candidates:
                try:
                    cls = autoclass(name)
                    activity = getattr(cls, "mActivity", None)
                    if activity is not None:
                        return activity
                except Exception as exc:
                    last_error = exc
        except Exception as exc:
            last_error = exc
        raise RuntimeError(f"Android Activity is unavailable: {last_error}")

    def _ensure_location_permission(self):
        """Prepare location access without using the Pydroid runtime-permission API.

        Pydroid 3 is a normal Python environment and does not provide the
        python-for-android ``android.permissions`` request flow.  In Pydroid,
        the app therefore does not open or simulate a permission dialog; it
        starts the normal Plyer GPS provider directly.  The packaged APK still
        uses the Android permission API because Buildozer declares the location
        permissions there.
        """
        if platform != "android":
            return True

        # Packaged python-for-android APK: request permission only here.
        try:
            from android.permissions import request_permissions, check_permission, Permission
            fine = bool(check_permission(Permission.ACCESS_FINE_LOCATION))
            coarse = bool(check_permission(Permission.ACCESS_COARSE_LOCATION))
            if fine or coarse:
                return True
            self._checkin_pending = True
            request_permissions(
                [Permission.ACCESS_FINE_LOCATION, Permission.ACCESS_COARSE_LOCATION],
                self._on_location_permission_result
            )
            return False
        except ImportError:
            # Pydroid 3: deliberately do not call Android requestPermissions().
            logger.info("Pydroid mode: skipping runtime location-permission request; using Plyer GPS.")
            return True
        except Exception:
            logger.exception("Packaged APK location-permission request failed.")
            self._show_gps_failure("Android Location permission could not be initialized.")
            return False

    @mainthread
    def _on_location_permission_result(self, permissions, grants):
        try:
            granted = any(bool(g) for g in grants)
        except Exception:
            granted = False
        if granted and self._checkin_pending:
            self._start_live_checkin_gps()
        elif self._checkin_pending:
            self._checkin_pending = False
            if self.dashboard_screen:
                self._show_gps_failure(
                    "Location permission is required. Allow Precise Location for ROHI Attendance, then press Check-In again."
                )
            try:
                self.dashboard_screen.ids.check_in_btn.disabled = False
            except Exception:
                pass

    def _cancel_gps_timeout(self):
        if self._gps_timeout_event is not None:
            try:
                self._gps_timeout_event.cancel()
            except Exception:
                pass
            self._gps_timeout_event = None

    def _start_android_location_polling(self):
        """Poll Android last-known location as a reliability fallback.

        Some Android/Pydroid builds successfully register LocationManager callbacks
        but never deliver the Java listener callback. Polling getLastKnownLocation
        still receives the provider's updated device fix and avoids the intermittent
        "no fresh GPS fix" failure seen on those builds.
        """
        if platform != "android":
            return
        try:
            from jnius import autoclass
            Context = autoclass("android.content.Context")
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            manager = PythonActivity.mActivity.getSystemService(Context.LOCATION_SERVICE)
            LocationManager = autoclass("android.location.LocationManager")
            providers = []
            for provider in (LocationManager.GPS_PROVIDER, LocationManager.NETWORK_PROVIDER):
                try:
                    if manager.isProviderEnabled(provider):
                        providers.append(provider)
                except Exception:
                    pass
            if not providers:
                return

            old_stop = self._gps_poll_stop
            if old_stop is not None:
                try:
                    old_stop.set()
                except Exception:
                    pass
            stop_event = threading.Event()
            self._gps_poll_stop = stop_event

            def worker():
                deadline = time.time() + 32
                last_signature = None
                while not stop_event.is_set() and time.time() < deadline and self._checkin_pending:
                    best = None
                    now_ms = int(time.time() * 1000)
                    for provider in providers:
                        try:
                            loc = manager.getLastKnownLocation(provider)
                            if loc is None:
                                continue
                            lat = float(loc.getLatitude())
                            lon = float(loc.getLongitude())
                            accuracy = float(loc.getAccuracy()) if loc.hasAccuracy() else 9999.0
                            loc_time = int(loc.getTime())
                            # Some Pydroid/Android providers return 0 for the timestamp
                            # even though the coordinate itself is valid. Treat timestamp 0
                            # as unknown age, not as an ancient location.
                            age_s = max(0.0, (now_ms - loc_time) / 1000.0) if loc_time > 0 else 0.0
                            if accuracy <= 150.0 and (loc_time <= 0 or age_s <= 600.0):
                                candidate = (accuracy, age_s, lat, lon)
                                if best is None or candidate[:2] < best[:2]:
                                    best = candidate
                        except Exception:
                            continue
                    if best is not None:
                        accuracy, age_s, lat, lon = best
                        signature = (round(lat, 6), round(lon, 6), round(accuracy, 1))
                        if signature != last_signature:
                            last_signature = signature
                            logger.info("Android GPS poll fix: %.7f, %.7f accuracy=%.1fm age=%.1fs", lat, lon, accuracy, age_s)
                            Clock.schedule_once(
                                lambda dt, la=lat, lo=lon, ac=accuracy: self._on_android_location(la, lo, ac),
                                0
                            )
                            return
                    time.sleep(1.0)

            threading.Thread(target=worker, daemon=True).start()
        except Exception:
            logger.exception("Android GPS polling fallback failed")

    def _start_live_checkin_gps(self):
        """Start Plyer GPS and an Android LocationManager fallback.

        Some Android/Pydroid combinations do not deliver a Plyer callback even
        though Location is enabled. The Android fallback asks LocationManager
        for a real device fix. A check-in is never approved from the office
        target coordinate itself.
        """
        self._cancel_gps_timeout()
        started = False
        try:
            gps.configure(on_location=self._on_checkin_gps, on_status=self.gps_status)
            gps.start(minTime=500, minDistance=0)
            started = True
            logger.info("Plyer GPS started for Check-In.")
        except Exception:
            logger.exception("Plyer GPS start failed; Android LocationManager fallback will be attempted.")

        if platform == "android":
            # LocationManager/pyjnius fallback is used only in the packaged
            # python-for-android APK. Pydroid 3 stays on Plyer GPS and never
            # attempts a runtime permission request through pyjnius.
            try:
                from android.permissions import check_permission, Permission
                fine = bool(check_permission(Permission.ACCESS_FINE_LOCATION))
                coarse = bool(check_permission(Permission.ACCESS_COARSE_LOCATION))
            except ImportError:
                fine = coarse = False
            except Exception:
                fine = coarse = False

            if fine or coarse:
                try:
                    self._start_android_location_fallback()
                    self._start_android_location_polling()
                    started = True
                except Exception:
                    logger.exception("Android LocationManager fallback failed.")

        if not started:
            self._checkin_pending = False
            self._show_gps_failure(
                "GPS could not be started. Turn ON Location/GPS and make sure the device allows Pydroid/ROHI to use location, then try again."
            )
            try:
                self.dashboard_screen.ids.check_in_btn.disabled = False
            except Exception:
                pass
            return

        # Give both providers enough time to obtain a satellite/network fix.
        self._gps_timeout_event = Clock.schedule_once(self._checkin_gps_timeout, 30)

    def _start_android_location_fallback(self):
        """Request a real Android location fix when Plyer is unreliable."""
        from jnius import autoclass, PythonJavaClass, java_method
        LocationManager = autoclass('android.location.LocationManager')
        Looper = autoclass('android.os.Looper')
        Context = autoclass('android.content.Context')

        app_context = None
        try:
            app_context = self._get_android_activity()
        except Exception:
            try:
                app_context = self._android_context
            except Exception:
                pass
        if app_context is None:
            raise RuntimeError("Android activity context unavailable")

        manager = app_context.getSystemService(Context.LOCATION_SERVICE)
        if manager is None:
            raise RuntimeError("Android LocationManager unavailable")

        outer = self

        class LocationListener(PythonJavaClass):
            __javainterfaces__ = ['android/location/LocationListener']
            __javacontext__ = 'app'

            @java_method('(Landroid/location/Location;)V')
            def onLocationChanged(self, location):
                try:
                    if location is not None:
                        lat = float(location.getLatitude())
                        lon = float(location.getLongitude())
                        accuracy = float(location.getAccuracy()) if location.hasAccuracy() else None
                        logger.info("Android LocationManager fix: %.7f, %.7f accuracy=%s", lat, lon, accuracy)
                        outer._on_android_location(lat, lon, accuracy)
                except Exception:
                    logger.exception("Android location callback failed")

            @java_method('(Ljava/lang/String;I)V')
            def onStatusChanged(self, provider, status):
                pass

            @java_method('(Ljava/lang/String;)V')
            def onProviderEnabled(self, provider):
                logger.info("Android location provider enabled: %s", provider)

            @java_method('(Ljava/lang/String;)V')
            def onProviderDisabled(self, provider):
                logger.warning("Android location provider disabled: %s", provider)

        listener = LocationListener()
        self._android_location_listener = listener

        providers = []
        try:
            if manager.isProviderEnabled(LocationManager.GPS_PROVIDER):
                providers.append(LocationManager.GPS_PROVIDER)
        except Exception:
            pass
        try:
            if manager.isProviderEnabled(LocationManager.NETWORK_PROVIDER):
                providers.append(LocationManager.NETWORK_PROVIDER)
        except Exception:
            pass
        if not providers:
            raise RuntimeError("No Android location provider is enabled")

        # A recent last-known fix is useful as a fast fallback, but only when
        # it is recent and reasonably accurate. It is still the phone's GPS,
        # never the office coordinate. We do NOT accept the office coordinate.
        now_ms = int(time.time() * 1000)
        best_last = None
        for provider in providers:
            try:
                last = manager.getLastKnownLocation(provider)
                if last is not None:
                    last_time = int(last.getTime())
                    age_ms = max(0, now_ms - last_time) if last_time > 0 else 0
                    accuracy = float(last.getAccuracy()) if last.hasAccuracy() else 9999.0
                    logger.info("Android last-known %s: age=%ss accuracy=%sm timestamp=%s", provider, age_ms / 1000.0, accuracy, last_time)
                    if (last_time <= 0 or age_ms <= 900000) and accuracy <= 150.0:
                        candidate = (age_ms, accuracy, float(last.getLatitude()), float(last.getLongitude()))
                        if best_last is None or candidate[:2] < best_last[:2]:
                            best_last = candidate
            except Exception:
                logger.exception("Could not read Android last-known location from %s", provider)

        # First request live updates from every enabled provider. This is more
        # reliable on Android/Pydroid than requestSingleUpdate, which can fail
        # silently on some devices. The first good fix finalizes Check-In.
        requested = False
        for provider in providers:
            try:
                manager.requestLocationUpdates(provider, 1000, 0.0, listener, Looper.getMainLooper())
                requested = True
                logger.info("Android live location updates requested from %s", provider)
            except Exception:
                logger.exception("Could not request live Android location updates from %s", provider)

        # Use a recent phone fix immediately when available. The live listener and
        # polling fallback continue in parallel, but this removes the Android/Pydroid
        # race where callbacks never arrive even though LocationManager has a valid fix.
        if best_last is not None:
            _, accuracy, lat, lon = best_last
            logger.info("Using recent Android phone location immediately: age/accuracy candidate=%s", best_last[:2])
            self._on_android_location(lat, lon, accuracy)
            return
        if not requested:
            raise RuntimeError("Android could not register any live location provider")

    @mainthread
    def _on_android_location(self, lat, lon, accuracy=None):
        if not self._checkin_pending:
            return
        # Do not accept a wildly inaccurate location as proof of being at the office.
        if accuracy is not None and accuracy > 150:
            logger.warning("Ignoring inaccurate Android location: accuracy=%.1fm", accuracy)
            return
        self._checkin_gps_fix = (float(lat), float(lon))
        self.current_location = f"{float(lat):.6f}° N, {float(lon):.6f}° E"
        logger.info("Fresh Android GPS accepted for Check-In: %s accuracy=%s", self.current_location, accuracy)
        self._cancel_gps_timeout()
        try:
            gps.stop()
        except Exception:
            pass
        self._stop_android_location_updates()
        self._finalize_check_in()

    def _stop_android_location_updates(self):
        """Stop Android LocationManager callbacks and polling after a fix or timeout."""
        try:
            if getattr(self, "_gps_poll_stop", None) is not None:
                self._gps_poll_stop.set()
                self._gps_poll_stop = None
        except Exception:
            pass
        try:
            if platform != "android" or not getattr(self, "_android_location_listener", None):
                return
            from jnius import autoclass
            Context = autoclass("android.content.Context")
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            manager = PythonActivity.mActivity.getSystemService(Context.LOCATION_SERVICE)
            if manager is not None:
                manager.removeUpdates(self._android_location_listener)
                logger.info("Android live location updates stopped.")
        except Exception:
            logger.exception("Could not stop Android location updates")
        finally:
            self._android_location_listener = None

    def _check_geofence(self, current_gps=None):
        """Compares a device location (defaults to the live self.current_location)
        against the staff member's office GPS coordinate, using that staff
        member's specific office radius (falls back to the default
        GEOFENCE_RADIUS_METERS if their State Office isn't in the OFFICES table).
        Returns (within_range: bool, distance_meters: float or None).

        IMPORTANT: for any staff member whose State Office is a known key in
        OFFICES, that live table is always the source of truth for the
        coordinate - NOT the gps_coordinate value stored on their record at
        registration time. This means correcting a coordinate in OFFICES
        (this file) takes effect immediately for every staff member at that
        office, without needing to edit/re-register each person. The stored
        registration value is only used as a fallback for staff whose office
        isn't in OFFICES."""
        if current_gps is None:
            current_gps = self.current_location

        radius = self.GEOFENCE_RADIUS_METERS
        state_office = str(self.current_user[13]).strip() if self.current_user and len(self.current_user) > 13 and self.current_user[13] else None
        office = OFFICES.get(state_office) if state_office else None

        if office:
            # Fixed ROHI office target; never replace the live phone GPS with this value.
            registered = f"{office['latitude']:.7f}, {office['longitude']:.7f}"
            radius = float(office.get("radius") or radius)
        elif self.current_user and len(self.current_user) > 26 and self.current_user[26]:
            registered = str(self.current_user[26])
        else:
            registered = self.static_gps

        distance = self._distance_meters(registered, current_gps)
        if distance is None:
            # Can't verify -> fail safe by rejecting the check-in.
            return False, None
        logger.info(
            f"Geofence check: office='{state_office}', registered='{registered}', "
            f"current='{current_gps}', distance={distance:.1f}m, radius={radius}m"
        )
        return distance <= radius, distance

    @staticmethod
    def _reset_threshold(dt):
        """Returns the next 6:00 PM boundary strictly after dt.
        Used so a Check-In/Check-Out stays static for the rest of the day
        and only clears at 6:00 PM to allow the next morning's check-in."""
        threshold = dt.replace(hour=18, minute=0, second=0, microsecond=0)
        if dt >= threshold:
            threshold += timedelta(days=1)
        return threshold

    def _reset_attendance_state(self, dash_ids=None):
        """Clears the static Check-In/Out session state (called at the 6PM boundary)."""
        self.checked_in = False
        self.check_in_datetime = None
        self.check_out_datetime = None
        self.late_duration_str = "On Time"
        if dash_ids is None and getattr(self, 'dashboard_screen', None):
            dash_ids = self.dashboard_screen.ids
        if not dash_ids:
            return
        if hasattr(dash_ids, 'clock_in_time_label'):
            dash_ids.clock_in_time_label.text = "Not Checked In"
        if hasattr(dash_ids, 'clock_out_time_label'):
            dash_ids.clock_out_time_label.text = "Not Checked Out"
        if hasattr(dash_ids, 'punctuality_status'):
            today = datetime.now()
            if self._is_attendance_working_day(today):
                dash_ids.punctuality_status.text = "Absent (Check-In Required)"
                dash_ids.punctuality_status.text_color = (0.8, 0.1, 0.1, 1)
            elif today.weekday() == 4:
                dash_ids.punctuality_status.text = "Work From Home (Friday)"
                dash_ids.punctuality_status.text_color = (0.2, 0.4, 0.8, 1)
            else:
                dash_ids.punctuality_status.text = "Non-Working Day"
                dash_ids.punctuality_status.text_color = (0.4, 0.4, 0.4, 1)
        if hasattr(dash_ids, 'hours_late_label'):
            dash_ids.hours_late_label.text = "Late by: 0m (On Time)"
        if hasattr(dash_ids, 'trend_today_status'):
            dash_ids.trend_today_status.text = "Pending"
            dash_ids.trend_today_status.text_color = (0.5, 0.5, 0.5, 1)
        if hasattr(dash_ids, 'check_in_btn'):
            dash_ids.check_in_btn.disabled = False
        if hasattr(dash_ids, 'check_out_btn'):
            dash_ids.check_out_btn.disabled = True
        if hasattr(dash_ids, 'geofence_note_label'):
            dash_ids.geofence_note_label.text = "Not Checked In Yet"
            dash_ids.geofence_note_label.text_color = (0.5, 0.5, 0.5, 1)
        if hasattr(dash_ids, 'geofence_icon'):
            dash_ids.geofence_icon.icon = "map-marker-question"
            dash_ids.geofence_icon.text_color = (0.5, 0.5, 0.5, 1)
        if hasattr(dash_ids, 'geofence_status_card'):
            dash_ids.geofence_status_card.md_bg_color = (0.95, 0.95, 0.95, 1)
        if hasattr(dash_ids, 'current_gps_label'):
            dash_ids.current_gps_label.text = "Not Checked In Yet"
        if hasattr(dash_ids, 'checkout_geofence_note_label'):
            dash_ids.checkout_geofence_note_label.text = "Not Checked Out Yet"
            dash_ids.checkout_geofence_note_label.text_color = (0.5, 0.5, 0.5, 1)
        if hasattr(dash_ids, 'checkout_geofence_icon'):
            dash_ids.checkout_geofence_icon.icon = "map-marker-question"
            dash_ids.checkout_geofence_icon.text_color = (0.5, 0.5, 0.5, 1)
        if hasattr(dash_ids, 'checkout_geofence_status_card'):
            dash_ids.checkout_geofence_status_card.md_bg_color = (0.95, 0.95, 0.95, 1)
        if hasattr(dash_ids, 'checkout_gps_label'):
            dash_ids.checkout_gps_label.text = "Not Checked Out Yet"

    # -----------------------------
    # 24-Hour Static Check-In/Out State Engine
    # -----------------------------
    def verify_existing_checkin(self, email):
        """Restores static check-in/out and attendance status if within 24 hours of last record."""
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        if not os.path.exists(db_path):
            return

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT check_in_time, check_out_time, late_duration, attendance_status, gps_location, check_out_gps_location FROM attendance 
            WHERE email=? ORDER BY id DESC LIMIT 1
        ''', (email,))
        record = cursor.fetchone()
        conn.close()

        dash_ids = self.dashboard_screen.ids

        if record:
            last_check_in_str = record[0]
            last_check_out_str = record[1]
            last_check_in_gps = record[4] if len(record) > 4 else None
            last_check_out_gps = record[5] if len(record) > 5 else None
            
            if last_check_in_str and last_check_in_str != "":
                last_check_in = datetime.strptime(last_check_in_str, "%Y-%m-%d %H:%M:%S")
            else:
                last_check_in = None

            if last_check_out_str and last_check_out_str != "":
                last_check_out = datetime.strptime(last_check_out_str, "%Y-%m-%d %H:%M:%S")
            else:
                last_check_out = None
            
            now = datetime.now()

            # Valid until the next 6:00 PM boundary after check-in; after that,
            # the session resets so the staff member can check in again the next morning.
            if last_check_in and now < self._reset_threshold(last_check_in):
                self.checked_in = True
                self.check_in_datetime = last_check_in
                self.check_out_datetime = last_check_out
                self.late_duration_str = record[2] or "On Time"
                
                if hasattr(dash_ids, 'clock_in_time_label'):
                    dash_ids.clock_in_time_label.text = last_check_in.strftime('%I:%M %p')
                if hasattr(dash_ids, 'check_in_btn'):
                    dash_ids.check_in_btn.disabled = True

                # Restore the Check-In GPS/geofence status box so it doesn't look like
                # nothing was captured just because the app was closed and reopened.
                if last_check_in_gps:
                    self.current_location = last_check_in_gps
                    self._last_checkin_gps = last_check_in_gps
                    within_range, distance = self._check_geofence(last_check_in_gps)
                    if hasattr(dash_ids, 'current_gps_label'):
                        dash_ids.current_gps_label.text = last_check_in_gps
                    note = (f"Captured Within Office ({distance:.0f} m)" if within_range and distance is not None
                            else f"Not Within Office Range ({distance:.0f} m away)" if distance is not None
                            else "Captured")
                    color = (0.1, 0.5, 0.15, 1) if within_range else (0.8, 0.1, 0.1, 1)
                    if hasattr(dash_ids, 'geofence_note_label'):
                        dash_ids.geofence_note_label.text = note
                        dash_ids.geofence_note_label.text_color = color
                    if hasattr(dash_ids, 'geofence_icon'):
                        dash_ids.geofence_icon.icon = "map-marker-check" if within_range else "map-marker-alert"
                        dash_ids.geofence_icon.text_color = color
                    if hasattr(dash_ids, 'geofence_status_card'):
                        dash_ids.geofence_status_card.md_bg_color = (
                            (0.88, 0.95, 0.88, 1) if within_range else (0.97, 0.88, 0.88, 1)
                        )

                if last_check_out:
                    if hasattr(dash_ids, 'clock_out_time_label'):
                        dash_ids.clock_out_time_label.text = last_check_out.strftime('%I:%M %p')
                    # Already checked out today -> keep the button locked and restore
                    # its GPS/geofence status box too (this was previously left blank,
                    # which made a re-opened app look like Check-Out never registered).
                    if hasattr(dash_ids, 'check_out_btn'):
                        dash_ids.check_out_btn.disabled = True
                    if last_check_out_gps:
                        out_within, out_distance = self._check_geofence(last_check_out_gps)
                        if hasattr(dash_ids, 'checkout_gps_label'):
                            dash_ids.checkout_gps_label.text = last_check_out_gps
                        out_note = (f"Captured Within Office ({out_distance:.0f} m)" if out_within and out_distance is not None
                                    else f"Not Within Office Range ({out_distance:.0f} m away)" if out_distance is not None
                                    else "Captured")
                        out_color = (0.1, 0.5, 0.15, 1) if out_within else (0.8, 0.1, 0.1, 1)
                        if hasattr(dash_ids, 'checkout_geofence_note_label'):
                            dash_ids.checkout_geofence_note_label.text = out_note
                            dash_ids.checkout_geofence_note_label.text_color = out_color
                        if hasattr(dash_ids, 'checkout_geofence_icon'):
                            dash_ids.checkout_geofence_icon.icon = "map-marker-check" if out_within else "map-marker-alert"
                            dash_ids.checkout_geofence_icon.text_color = out_color
                        if hasattr(dash_ids, 'checkout_geofence_status_card'):
                            dash_ids.checkout_geofence_status_card.md_bg_color = (
                                (0.88, 0.95, 0.88, 1) if out_within else (0.97, 0.88, 0.88, 1)
                            )
                else:
                    if hasattr(dash_ids, 'clock_out_time_label'):
                        dash_ids.clock_out_time_label.text = "Pending"
                    # Not checked out yet -> make sure the button is actually usable
                    # (it may have been left disabled by a previous session/crash).
                    if hasattr(dash_ids, 'check_out_btn'):
                        dash_ids.check_out_btn.disabled = False

                status_text = record[3] or "Present (On Time)"
                if hasattr(dash_ids, 'punctuality_status'):
                    dash_ids.punctuality_status.text = status_text
                    if "Late" in status_text:
                        dash_ids.punctuality_status.text_color = (0.9, 0.5, 0, 1) # Yellow
                    elif "Present" in status_text:
                        dash_ids.punctuality_status.text_color = (0.1, 0.6, 0.2, 1) # Green
                    else:
                        dash_ids.punctuality_status.text_color = (0.8, 0.1, 0.1, 1) # Red

                if hasattr(dash_ids, 'hours_late_label'):
                    dash_ids.hours_late_label.text = (
                        "Late by: 0m (On Time)" if self.late_duration_str == "On Time"
                        else f"Late by: {self.late_duration_str}"
                    )
                return

        # No valid same-day check-in found (new staff member, or the previous
        # session already rolled past the 6PM boundary) - fully reset every
        # dashboard label via _reset_attendance_state() so nothing is left
        # showing a stale value or the raw kv placeholder text. (Previously
        # this duplicated only part of that reset - e.g. hours_late_label,
        # trend_today_status and the geofence boxes were never touched here,
        # so on a normal fresh-day login they were stuck on whatever text was
        # hardcoded in dashboard.kv, such as "Hrs Late: 00:00 Mins".)
        self._reset_attendance_state(dash_ids)

    def update_dashboard_time(self, dt=None):
        """Updates dynamic clocks and handles daily reset/expiration at 8:00 AM cycle."""
        try:
            now = datetime.now()
            formatted_time = now.strftime("%A, %d %B %Y | %I:%M:%S %p")
            dash_ids = self.dashboard_screen.ids
            
            if hasattr(dash_ids, 'live_time_label'):
                dash_ids.live_time_label.text = formatted_time

            # Handle automatic expiration / reset once the 6:00 PM boundary passes,
            # clearing the static Check-In/Out state for the next morning.
            if self.checked_in and self.check_in_datetime:
                if now >= self._reset_threshold(self.check_in_datetime):
                    self._reset_attendance_state(dash_ids)

            if hasattr(dash_ids, 'punctuality_status') and not self.checked_in:
                if self._is_attendance_working_day(now):
                    work_start_time = now.replace(hour=8, minute=0, second=0, microsecond=0)
                    if now > work_start_time:
                        dash_ids.punctuality_status.text = "Absent (Check-In Required)"
                        dash_ids.punctuality_status.text_color = (0.8, 0.1, 0.1, 1)
                elif now.weekday() == 4:
                    dash_ids.punctuality_status.text = "Work From Home (Friday)"
                    dash_ids.punctuality_status.text_color = (0.2, 0.4, 0.8, 1)
                else:
                    dash_ids.punctuality_status.text = "Non-Working Day"
                    dash_ids.punctuality_status.text_color = (0.4, 0.4, 0.4, 1)
        except Exception:
            logger.exception("Error in update_dashboard_time:")

    def update_dashboard_metrics(self):
        """Calculates working hours today, monthly working days, and real
        present/absent/late/punctuality figures from the attendance table."""
        try:
            now = datetime.now()
            dash_ids = self.dashboard_screen.ids

            weekday = now.weekday()
            if hasattr(dash_ids, 'expected_hours_today'):
                if weekday < 4:
                    dash_ids.expected_hours_today.text = "9 Hours"
                elif weekday == 4:
                    dash_ids.expected_hours_today.text = "WFH - Attendance Not Captured"
                else:
                    dash_ids.expected_hours_today.text = "0 Hours (Weekend)"

            year, month = now.year, now.month
            num_days = calendar.monthrange(year, month)[1]
            # Attendance working days are Monday-Thursday only. Friday is WFH
            # and is deliberately excluded from attendance-day totals.
            work_days = sum(1 for day in range(1, num_days + 1) if calendar.weekday(year, month, day) in (0, 1, 2, 3))
            work_days_elapsed = sum(
                1 for day in range(1, now.day + 1) if calendar.weekday(year, month, day) in (0, 1, 2, 3)
            )

            if hasattr(dash_ids, 'total_work_days'):
                dash_ids.total_work_days.text = f"{work_days} Days"

            days_present = 0
            days_late = 0

            email_val = str(self.current_user[20]) if self.current_user and len(self.current_user) > 20 else None
            if email_val:
                db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
                if os.path.exists(db_path):
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    month_prefix = f"{year:04d}-{month:02d}-"
                    cursor.execute(
                        '''
                        SELECT check_in_time, attendance_status FROM attendance
                        WHERE email = ? AND check_in_time LIKE ?
                        ''',
                        (email_val, f"{month_prefix}%"),
                    )
                    rows = cursor.fetchall()
                    conn.close()

                    present_days_seen = set()
                    for check_in_time, status in rows:
                        if not check_in_time:
                            continue
                        day_key = check_in_time[:10]
                        present_days_seen.add(day_key)
                        if status and "Late" in status:
                            days_late += 1
                    days_present = len(present_days_seen)

            # Missing attendance rows are not automatic absences.
            days_absent = 0
            if email_val:
                db_path = os.path.join(APP_DIR, "attendance.db")
                if os.path.exists(db_path):
                    conn = sqlite3.connect(db_path)
                    try:
                        cursor = conn.cursor()
                        month_prefix = f"{year:04d}-{month:02d}-"
                        cursor.execute("SELECT check_in_time, attendance_status FROM attendance WHERE email = ? AND check_in_time LIKE ?", (email_val, f"{month_prefix}%"))
                        days_absent = len({r[0][:10] for r in cursor.fetchall() if r[0] and r[1] and "absent" in str(r[1]).lower()})
                    finally: conn.close()
            absence_percentage = (days_absent / work_days_elapsed * 100) if work_days_elapsed > 0 else 0.0
            punctuality_rate = ((days_present - days_late) / days_present * 100) if days_present > 0 else 0.0

            if hasattr(dash_ids, 'absence_rate'):
                dash_ids.absence_rate.text = f"{absence_percentage:.1f}%"
            if hasattr(dash_ids, 'total_days_present'):
                dash_ids.total_days_present.text = f"{days_present} Days"
            if hasattr(dash_ids, 'total_days_absent'):
                dash_ids.total_days_absent.text = f"{days_absent} Days"
            if hasattr(dash_ids, 'card_absent_count'):
                dash_ids.card_absent_count.text = f"{days_absent} Days"
            if hasattr(dash_ids, 'card_punctuality_rate'):
                dash_ids.card_punctuality_rate.text = f"{punctuality_rate:.0f}%"
            if hasattr(dash_ids, 'card_late_count'):
                dash_ids.card_late_count.text = f"{days_late} Times"

        except Exception:
            logger.exception("Error calculating dashboard metrics:")

    # -----------------------------
    # Attendance Working-Day Rules
    # -----------------------------
    @staticmethod
    def _is_attendance_working_day(date_obj=None):
        """Attendance capture is permitted Monday-Thursday only.
        Friday is official Work From Home and remains in the timesheet; it is
        not a check-in/check-out day. Saturday and Sunday are non-working days.
        """
        date_obj = date_obj or datetime.now()
        return date_obj.weekday() in (0, 1, 2, 3)

    def _block_attendance_non_working_day(self, action_name='Attendance'):
        now = datetime.now()
        if self._is_attendance_working_day(now):
            return False
        if now.weekday() == 4:
            title = f"{action_name} Unavailable - Friday WFH"
            message = ("Friday is Work From Home (WFH).\n\n"
                       "Check-In and Check-Out are not captured on Friday. "
                       "Friday remains included in the Timesheet.")
        else:
            title = f"{action_name} Unavailable"
            message = ("Attendance can only be captured Monday to Thursday.\n\n"
                       "Friday is Work From Home, while Saturday and Sunday are non-working days.")
        dialog = MDDialog(title=title, text=message,
                          buttons=[MDFlatButton(text="OK", on_release=lambda x: dialog.dismiss())])
        dialog.open()
        logger.info("%s blocked on non-attendance day: %s", action_name, now.strftime('%A'))
        return True

    # -----------------------------
    # Attendance Reminder Engine
    # -----------------------------
    def _reminder_reset_if_new_day(self, today):
        if self._reminder_state.get("date") != today:
            self._reminder_state = {"date": today, "checkin": set(), "checkout": set()}

    def _today_attendance_row(self):
        """Latest attendance row for the logged-in user for today, or None."""
        try:
            if not self.current_user or len(self.current_user) <= 20:
                return None
            email_val = str(self.current_user[20])
            today_prefix = datetime.now().strftime("%Y-%m-%d")
            db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
            if not os.path.exists(db_path):
                return None
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT check_in_time, check_out_time FROM attendance "
                "WHERE email = ? AND check_in_time LIKE ? ORDER BY id DESC LIMIT 1",
                (email_val, f"{today_prefix}%"),
            )
            row = cursor.fetchone()
            conn.close()
            return row
        except Exception:
            logger.exception("Failed to read today's attendance for reminders:")
            return None

    def _reminder_tick(self, *args):
        """Runs every 30s. Fires the on-time notification plus up to
        REMINDER_REPEAT_COUNT nudges (every REMINDER_REPEAT_MINUTES) for
        whichever of check-in/check-out is still outstanding. Automatically
        skips Friday (WFH) and Saturday/Sunday via _is_attendance_working_day,
        and stops nudging the moment the action is recorded."""
        try:
            now = datetime.now()
            if not self.current_user:
                return
            if not self._load_reminders_enabled():
                return
            if not self._is_attendance_working_day(now):
                return
            self._reminder_reset_if_new_day(now.date())

            row = self._today_attendance_row()
            checked_in = bool(row and row[0])
            checked_out = bool(row and row[1])

            self._maybe_fire_reminder(
                "checkin", now, REMINDER_CHECKIN_HOUR, REMINDER_CHECKIN_MINUTE,
                checked_in, CHECKIN_REMINDER_OFFSETS_MINUTES
            )
            self._maybe_fire_reminder(
                "checkout", now, REMINDER_CHECKOUT_HOUR, REMINDER_CHECKOUT_MINUTE,
                checked_out, CHECKOUT_REMINDER_OFFSETS_MINUTES
            )
        except Exception:
            logger.exception("Reminder tick failed:")

    def _maybe_fire_reminder(self, kind, now, hour, minute, already_done, offsets_minutes):
        if already_done:
            return
        base_time = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
        fired = self._reminder_state[kind]
        last_slot = len(offsets_minutes) - 1
        for slot, offset in enumerate(offsets_minutes):  # 0 = on-time, 1..N = nudges
            if slot in fired:
                continue
            slot_time = base_time + timedelta(minutes=offset)
            if now >= slot_time:
                fired.add(slot)
                self._send_reminder_notification(kind, slot, last_slot)

    @staticmethod
    def _send_reminder_notification(kind, slot, last_slot):
        label = "Check-In" if kind == "checkin" else "Check-Out"
        verb = "checked in" if kind == "checkin" else "checked out"
        if slot == 0:
            title = f"ROHI {label} Reminder"
            message = f"It's time to {label.lower()}. Open ROHI Attendance and tap {label}."
        else:
            title = f"ROHI {label} Reminder ({slot}/{last_slot})"
            message = f"You still haven't {verb} today. Please do so now."
        try:
            from plyer import notification as plyer_notification
            plyer_notification.notify(title=title, message=message, app_name="ROHI Attendance", timeout=15)
        except Exception:
            logger.exception("Failed to show %s reminder notification (slot %s):", kind, slot)
        logger.info("Reminder fired: %s slot %s", kind, slot)

    # -----------------------------
    # Silent Background Auto-Sync
    # -----------------------------
    def _auto_sync_tick(self, *args):
        try:
            config = pg_sync.load_config()
            if config.get("host"):
                threading.Thread(target=self._auto_sync_worker, args=(config,), daemon=True).start()
        except Exception:
            logger.exception("Auto-sync tick failed:")

        try:
            if gform_sync.is_configured():
                threading.Thread(target=self._gform_auto_sync_worker, daemon=True).start()
        except Exception:
            logger.exception("Google Form auto-sync tick failed:")

    @staticmethod
    def _auto_sync_worker(config):
        try:
            ok, message, counts = pg_sync.synchronize_now(config)
            logger.info("Auto-sync: ok=%s message=%s counts=%s", ok, message, counts)
        except Exception:
            logger.exception("Background auto-sync failed:")

    @staticmethod
    def _gform_auto_sync_worker():
        """Pushes every completed (checked-out) attendance row that hasn't
        reached the Google Form yet. Runs on the shared AUTO_SYNC_INTERVAL_SECONDS
        timer, and is also kicked off immediately after a check-out so today's
        row shows up in the Sheet right away instead of waiting for the timer."""
        try:
            config = gform_sync.load_config()
            rows = get_pending_gform_attendance()
            for row in rows:
                (att_id, email, check_in_time, check_out_time, gps_in, gps_out,
                 fullname, staff_number, department, section, position) = row
                date_str = (check_in_time or check_out_time or "")[:10]
                payload = {
                    "name": fullname or email or "",
                    "staff_id": staff_number or "",
                    "department": department or "",
                    "section": section or "",
                    "position": position or "",
                    "date": date_str,
                    "checkin": check_in_time or "",
                    "checkout": check_out_time or "",
                    "gps": gps_out or gps_in or "",
                }
                ok, message = gform_sync.submit_row(payload, config=config)
                if ok:
                    mark_gform_synced(att_id)
                    logger.info("Google Form sync: row id=%s submitted.", att_id)
                else:
                    logger.warning("Google Form sync: row id=%s failed: %s", att_id, message)
                    # Stop on first failure (likely offline) - the next timer
                    # tick will retry this and any later rows in order.
                    break
        except Exception:
            logger.exception("Google Form background sync failed:")

    # -----------------------------
    # Assigned Check In & Check Out Actions
    # -----------------------------
    def _show_gps_failure(self, message):
        try:
            dash_ids = self.dashboard_screen.ids
            if hasattr(dash_ids, 'geofence_note_label'):
                dash_ids.geofence_note_label.text = message
                dash_ids.geofence_note_label.text_color = (0.8, 0.1, 0.1, 1)
            if hasattr(dash_ids, 'checkout_geofence_note_label'):
                dash_ids.checkout_geofence_note_label.text = message
                dash_ids.checkout_geofence_note_label.text_color = (0.8, 0.1, 0.1, 1)
        except Exception:
            logger.exception("Could not display GPS failure message:")

    def check_in(self):
        """Captures a fresh GPS fix and validates it against the staff member's
        registered office coordinate before completing Check-In (geofencing)."""
        if self._block_attendance_non_working_day("Check-In"):
            return
        logger.info("Check-In action triggered; capturing GPS for geofence validation...")
        dash_ids = self.dashboard_screen.ids
        if self.checked_in:
            logger.info("User is already checked in and static lock is active.")
            if hasattr(dash_ids, 'geofence_note_label'):
                dash_ids.geofence_note_label.text = "✅ Already checked in today."
                dash_ids.geofence_note_label.text_color = (0.13, 0.40, 0.16, 1)
            return
        if not self.current_user:
            logger.warning("Check-In blocked: no authenticated user in session.")
            return

        # Immediate feedback so the (up to a few seconds) GPS wait doesn't look frozen.
        if hasattr(dash_ids, 'check_in_btn'):
            dash_ids.check_in_btn.disabled = True
        if hasattr(dash_ids, 'geofence_note_label'):
            dash_ids.geofence_note_label.text = "📍 Capturing current GPS location..."
            dash_ids.geofence_note_label.text_color = (0.4, 0.4, 0.4, 1)
        if hasattr(dash_ids, 'geofence_icon'):
            dash_ids.geofence_icon.icon = "crosshairs-gps"
            dash_ids.geofence_icon.text_color = (0.4, 0.4, 0.4, 1)

        self._checkin_pending = True
        self._checkin_gps_fix = None
        self.current_location = ""
        # Always obtain a fresh phone GPS fix. The office coordinate is only
        # the fixed geofence target; it is never substituted for the phone's
        # current coordinate.
        if not self._ensure_location_permission():
            return
        self._start_live_checkin_gps()

    def _on_checkin_gps(self, **kwargs):
        # NOTE: on Android this callback fires on plyer's location-listener
        # thread, not the Kivy UI thread. @mainthread schedules the actual
        # work for the next UI frame so widget updates are safe and reliable
        # (previously this ran inline and could silently fail to update the
        # dashboard, or intermittently drop the GPS fix, on real devices).
        lat = kwargs.get("lat")
        lon = kwargs.get("lon")
        self._handle_checkin_gps(lat, lon)

    @mainthread
    def _handle_checkin_gps(self, lat, lon):
        if not self._checkin_pending:
            return
        if lat is not None and lon is not None:
            self._checkin_gps_fix = (float(lat), float(lon))
            self.current_location = f"{float(lat):.6f}° N, {float(lon):.6f}° E"
            self._last_checkin_gps = self.current_location
            logger.info(f"Check-In fresh Plyer GPS captured: {self.current_location}")
            self._cancel_gps_timeout()
            try:
                gps.stop()
            except Exception:
                pass
            self._finalize_check_in()

    def _checkin_gps_timeout(self, dt):
        self._gps_timeout_event = None
        if self._checkin_pending:
            logger.warning("GPS fix timed out during check-in; denying because no fresh phone coordinate was captured.")
            try:
                gps.stop()
            except Exception:
                pass
            self._stop_android_location_updates()
            self._checkin_pending = False
            self._show_gps_failure("Check-In denied: no fresh phone GPS fix was received within 30 seconds. Confirm Android Location is ON and Precise Location is allowed for ROHI/Pydroid, then press CHECK IN again.")
            try:
                self.dashboard_screen.ids.check_in_btn.disabled = False
            except Exception:
                pass

    def _finalize_check_in(self):
        """Validates the geofence against the freshly captured GPS, then completes
        Check-In only if the staff member is within office range."""
        if not self._checkin_pending:
            return
        self._checkin_pending = False

        dash_ids = self.dashboard_screen.ids
        try:
            if self.checked_in:
                return

            now = datetime.now()
            if not self._checkin_gps_fix or not self.current_location:
                if hasattr(dash_ids, 'current_gps_label'):
                    dash_ids.current_gps_label.text = "GPS unavailable"
                if hasattr(dash_ids, 'geofence_note_label'):
                    dash_ids.geofence_note_label.text = "GPS unavailable - Check-In Denied"
                return
            within_range, distance = self._check_geofence(self.current_location)

            if hasattr(dash_ids, 'current_gps_label'):
                dash_ids.current_gps_label.text = self.current_location

            if not within_range:
                note = (f"Not Within Office Range ({distance:.0f} m away) - Check-In Denied"
                        if distance is not None else "Location Unavailable - Check-In Denied")
                if hasattr(dash_ids, 'geofence_note_label'):
                    dash_ids.geofence_note_label.text = note
                    dash_ids.geofence_note_label.text_color = (0.8, 0.1, 0.1, 1)
                if hasattr(dash_ids, 'geofence_icon'):
                    dash_ids.geofence_icon.icon = "map-marker-alert"
                    dash_ids.geofence_icon.text_color = (0.8, 0.1, 0.1, 1)
                if hasattr(dash_ids, 'geofence_status_card'):
                    dash_ids.geofence_status_card.md_bg_color = (0.97, 0.88, 0.88, 1)

                distance_text = f"You are currently about {distance:.0f} m away." if distance is not None else "Your location could not be verified."
                dialog = MDDialog(
                    title="Check-In Denied",
                    text=f"You must be within {self.GEOFENCE_RADIUS_METERS} m of the registered office GPS coordinate to check in. {distance_text}",
                    buttons=[MDFlatButton(text="OK", on_release=lambda x: dialog.dismiss())]
                )
                dialog.open()
                logger.warning(f"Check-in rejected: outside geofence (distance={distance}).")
                return

            # Within range -> proceed with Check-In
            self.checked_in = True
            self.check_in_datetime = now
            self.check_out_datetime = None

            # Punctuality & Late Hours Calculation (Target: 8:00 AM)
            work_start_time = now.replace(hour=8, minute=0, second=0, microsecond=0)
            if now > work_start_time:
                diff = now - work_start_time
                hours, remainder = divmod(diff.seconds, 3600)
                minutes = remainder // 60
                self.late_duration_str = f"{hours}h {minutes}m" if hours else f"{minutes}m"
                status_str = "Present but Late"
                text_color = (0.9, 0.5, 0, 1)  # Yellow
            else:
                self.late_duration_str = "On Time"
                status_str = "Present (On Time)"
                text_color = (0.1, 0.6, 0.2, 1)  # Green

            if hasattr(dash_ids, 'clock_in_time_label'):
                dash_ids.clock_in_time_label.text = now.strftime('%I:%M %p')
            if hasattr(dash_ids, 'clock_out_time_label'):
                dash_ids.clock_out_time_label.text = "Pending"
            if hasattr(dash_ids, 'punctuality_status'):
                dash_ids.punctuality_status.text = status_str
                dash_ids.punctuality_status.text_color = text_color
            if hasattr(dash_ids, 'hours_late_label'):
                dash_ids.hours_late_label.text = (
                    "Late by: 0m (On Time)" if self.late_duration_str == "On Time"
                    else f"Late by: {self.late_duration_str}"
                )
            if hasattr(dash_ids, 'trend_today_status'):
                dash_ids.trend_today_status.text = "⚠ Late" if status_str == "Present but Late" else "✓ On"
                dash_ids.trend_today_status.text_color = text_color

            note = f"Captured Within Office ({distance:.0f} m)" if distance is not None else "Captured Within Office"
            if hasattr(dash_ids, 'geofence_note_label'):
                dash_ids.geofence_note_label.text = note
                dash_ids.geofence_note_label.text_color = (0.1, 0.5, 0.15, 1)
            if hasattr(dash_ids, 'geofence_icon'):
                dash_ids.geofence_icon.icon = "map-marker-check"
                dash_ids.geofence_icon.text_color = (0.1, 0.6, 0.2, 1)
            if hasattr(dash_ids, 'geofence_status_card'):
                dash_ids.geofence_status_card.md_bg_color = (0.88, 0.95, 0.88, 1)

            # Save attendance log to database with exact day-to-day exact time-in and time-out log support
            email_val = str(self.current_user[20]) if self.current_user and len(self.current_user) > 20 else "user@rohi.org"
            db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO attendance (email, check_in_time, check_out_time, late_duration, attendance_status, gps_location, synced)
                VALUES (?, ?, ?, ?, ?, ?, 0)
            ''', (email_val, now.strftime("%Y-%m-%d %H:%M:%S"), "", self.late_duration_str, status_str, self.current_location))
            conn.commit()
            conn.close()

            # Kobo-style immediate submission: send the attendance row as data
            # to the configured endpoint, without opening Google or generating
            # an Excel file for every check-in.
            threading.Thread(
                target=self._submit_attendance_to_endpoint,
                kwargs={"check_in": now.strftime("%Y-%m-%d %H:%M:%S"),
                        "checkin_gps": self.current_location},
                daemon=True,
            ).start()

            self.update_dashboard_metrics()
            logger.info("Staff successfully checked in.")
        except Exception:
            logger.exception("Error executing check_in:")
        finally:
            if hasattr(dash_ids, 'check_in_btn'):
                dash_ids.check_in_btn.disabled = self.checked_in
            # Check-Out only becomes usable once Check-In has actually succeeded.
            if hasattr(dash_ids, 'check_out_btn'):
                dash_ids.check_out_btn.disabled = not self.checked_in

    def check_out(self):
        """Captures a fresh GPS fix at the moment Check-Out is pressed, and
        evaluates it against the registered office coordinate (informational -
        unlike Check-In, Check-Out is not blocked by geofencing, only labeled)."""
        if self._block_attendance_non_working_day("Check-Out"):
            return
        logger.info("Check-Out action triggered; capturing GPS...")
        dash_ids = self.dashboard_screen.ids
        if not self.checked_in:
            logger.info("User is not checked in yet.")
            if hasattr(dash_ids, 'checkout_geofence_note_label'):
                dash_ids.checkout_geofence_note_label.text = "⚠️ You need to Check In first."
                dash_ids.checkout_geofence_note_label.text_color = (0.8, 0.1, 0.1, 1)
            return
        if self.check_out_datetime:
            logger.info("User already checked out; static lock is active.")
            if hasattr(dash_ids, 'checkout_geofence_note_label'):
                dash_ids.checkout_geofence_note_label.text = "✅ Already checked out today."
                dash_ids.checkout_geofence_note_label.text_color = (0.13, 0.40, 0.16, 1)
            return

        if hasattr(dash_ids, 'check_out_btn'):
            dash_ids.check_out_btn.disabled = True
        if hasattr(dash_ids, 'checkout_geofence_note_label'):
            dash_ids.checkout_geofence_note_label.text = "📍 Capturing current GPS location..."
            dash_ids.checkout_geofence_note_label.text_color = (0.4, 0.4, 0.4, 1)
        if hasattr(dash_ids, 'checkout_geofence_icon'):
            dash_ids.checkout_geofence_icon.icon = "crosshairs-gps"
            dash_ids.checkout_geofence_icon.text_color = (0.4, 0.4, 0.4, 1)

        self._checkout_pending = True
        self._checkout_gps_fix = None
        self._android_location_listener = None
        self._gps_poll_stop = None
        self._gps_timeout_event = None
        self.current_location = ""
        try:
            gps.configure(on_location=self._on_checkout_gps, on_status=self.gps_status)
            gps.start(minTime=1000, minDistance=1)
            Clock.schedule_once(self._checkout_gps_timeout, 10)
        except Exception:
            logger.exception("GPS start error during check-out; no fresh phone coordinate available.")
            self._checkout_pending = False
            self._show_gps_failure("Check-Out denied: GPS could not be started. Please enable Location and try again.")

    def _on_checkout_gps(self, **kwargs):
        # See _on_checkin_gps: marshal off the plyer callback thread and
        # onto the Kivy UI thread before touching any widgets or finalizing.
        lat = kwargs.get("lat")
        lon = kwargs.get("lon")
        self._handle_checkout_gps(lat, lon)

    @mainthread
    def _handle_checkout_gps(self, lat, lon):
        if lat is not None and lon is not None:
            self._checkout_gps_fix = (float(lat), float(lon))
            self.current_location = f"{float(lat):.6f}° N, {float(lon):.6f}° E"
            logger.info(f"Check-Out fresh GPS captured: {self.current_location}")
            try:
                gps.stop()
            except Exception:
                pass
            self._finalize_check_out()

    def _checkout_gps_timeout(self, dt):
        if getattr(self, '_checkout_pending', False):
            logger.warning("GPS fix timed out during check-out; no fresh phone coordinate available.")
            try:
                gps.stop()
            except Exception:
                pass
            self._checkout_pending = False
            self._show_gps_failure("Check-Out denied: the phone could not obtain a fresh GPS coordinate. Please enable Location and try again.")

    def _finalize_check_out(self):
        """Completes Check-Out using the freshly captured GPS: records the exact
        time + coordinate, and shows Within/Not Within Office status (informational)."""
        if not getattr(self, '_checkout_pending', False):
            return
        self._checkout_pending = False

        dash_ids = self.dashboard_screen.ids
        try:
            if not self.checked_in or self.check_out_datetime:
                return

            now = datetime.now()
            if not self._checkout_gps_fix or not self.current_location:
                return
            checkout_gps = self.current_location
            within_range, distance = self._check_geofence(checkout_gps)

            self.check_out_datetime = now

            if hasattr(dash_ids, 'clock_out_time_label'):
                dash_ids.clock_out_time_label.text = now.strftime('%I:%M %p')
            if hasattr(dash_ids, 'checkout_gps_label'):
                dash_ids.checkout_gps_label.text = checkout_gps

            if within_range:
                note = f"Captured Within Office ({distance:.0f} m)" if distance is not None else "Captured Within Office"
                color = (0.1, 0.5, 0.15, 1)
                icon = "map-marker-check"
                card_color = (0.88, 0.95, 0.88, 1)
            else:
                note = f"Not Within Office Range ({distance:.0f} m away)" if distance is not None else "Not Within Office Range"
                color = (0.8, 0.1, 0.1, 1)
                icon = "map-marker-alert"
                card_color = (0.97, 0.88, 0.88, 1)

            if hasattr(dash_ids, 'checkout_geofence_note_label'):
                dash_ids.checkout_geofence_note_label.text = note
                dash_ids.checkout_geofence_note_label.text_color = color
            if hasattr(dash_ids, 'checkout_geofence_icon'):
                dash_ids.checkout_geofence_icon.icon = icon
                dash_ids.checkout_geofence_icon.text_color = color
            if hasattr(dash_ids, 'checkout_geofence_status_card'):
                dash_ids.checkout_geofence_status_card.md_bg_color = card_color

            # Update database record with exact check_out_time AND the GPS
            # coordinate captured at that moment, for monthly report maintenance.
            email_val = str(self.current_user[20]) if self.current_user and len(self.current_user) > 20 else "user@rohi.org"
            db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            # Scope strictly to TODAY's check-in row (matching on the exact
            # check_in_datetime we hold in memory). Previously this matched
            # ANY row for this email with an empty check_out_time, which could
            # silently stamp the checkout onto a stale old row instead of
            # today's, making today's checkout look like it never happened.
            today_prefix = now.strftime("%Y-%m-%d")
            cursor.execute('''
                UPDATE attendance SET check_out_time = ?, check_out_gps_location = ?, synced = 0
                WHERE email = ? AND check_in_time = ?
            ''', (now.strftime("%Y-%m-%d %H:%M:%S"), checkout_gps, email_val,
                  self.check_in_datetime.strftime("%Y-%m-%d %H:%M:%S")))

            if cursor.rowcount == 0:
                # Fallback: match today's most recent still-open row for this email.
                cursor.execute('''
                    UPDATE attendance SET check_out_time = ?, check_out_gps_location = ?, synced = 0
                    WHERE id = (
                        SELECT id FROM attendance
                        WHERE email = ? AND check_in_time LIKE ?
                        AND (check_out_time IS NULL OR check_out_time = '')
                        ORDER BY id DESC LIMIT 1
                    )
                ''', (now.strftime("%Y-%m-%d %H:%M:%S"), checkout_gps, email_val, f"{today_prefix}%"))

            conn.commit()
            conn.close()

            # Submit the completed row immediately to the Attendance endpoint.
            threading.Thread(
                target=self._submit_attendance_to_endpoint,
                kwargs={"check_in": self.check_in_datetime.strftime("%Y-%m-%d %H:%M:%S") if self.check_in_datetime else "",
                        "check_out": now.strftime("%Y-%m-%d %H:%M:%S"),
                        "checkin_gps": getattr(self, "_last_checkin_gps", "") or "",
                        "checkout_gps": checkout_gps},
                daemon=True,
            ).start()

            self.update_dashboard_metrics()
            logger.info("Staff successfully checked out. Session stays static until the 6:00 PM reset.")

            # Fire an immediate Google Form push for today's now-complete row,
            # instead of waiting for the next 5-minute auto-sync tick.
            try:
                if gform_sync.is_configured():
                    threading.Thread(target=self._gform_auto_sync_worker, daemon=True).start()
            except Exception:
                logger.exception("Failed to kick off immediate Google Form sync after check-out:")
        except Exception:
            logger.exception("Error executing check_out:")
        finally:
            if hasattr(dash_ids, 'check_out_btn'):
                dash_ids.check_out_btn.disabled = bool(self.check_out_datetime)

    # -----------------------------
    # Dropdown & Picker Handlers
    # -----------------------------
    def open_sex_menu(self):
        self._dismiss_active_menu()
        options = ["Male", "Female"]
        items = [
            {"text": opt, "viewclass": "OneLineListItem", "on_release": lambda x=opt: self._set_field_text('sex', x)}
            for opt in options
        ]
        self.active_menu = MDDropdownMenu(caller=self.registration_screen.ids.sex, items=items, width_mult=4)
        self.active_menu.open()

    def open_state_office_menu(self):
        self._dismiss_active_menu()
        items = [
            {"text": off, "viewclass": "OneLineListItem", "on_release": lambda x=off: self._select_state_office(x)}
            for off in STATE_OFFICES
        ]
        self.active_menu = MDDropdownMenu(caller=self.registration_screen.ids.state_office, items=items, width_mult=4)
        self.active_menu.open()

    def _select_state_office(self, office_name):
        """Sets the State field AND auto-fills the (now static) GPS
        Coordinate used for geofencing from the fixed OFFICES lookup, instead
        of the device's live GPS. This coordinate does not change again after
        being set here. The registration form no longer shows a raw GPS
        Coordinate field, so the value is cached on the app and picked up by
        submit_staff() directly."""
        self._set_field_text('state_office', office_name)
        office = OFFICES.get(office_name)
        if office:
            coord_str = f"{office['latitude']}° N, {office['longitude']}° E"
            self._selected_office_gps_coordinate = coord_str
            if hasattr(self.registration_screen.ids, 'gps_coordinate'):
                self.registration_screen.ids.gps_coordinate.text = coord_str
            logger.info(f"GPS Coordinate auto-filled from '{office_name}': {coord_str}")

    def open_department_menu(self):
        self._dismiss_active_menu()
        departments = ["Programs", "HR & Operation", "Finance"]
        items = [
            {"text": dept, "viewclass": "OneLineListItem", "on_release": lambda x=dept: self._on_department_select(x)}
            for dept in departments
        ]
        self.active_menu = MDDropdownMenu(caller=self.registration_screen.ids.department, items=items, width_mult=4)
        self.active_menu.open()

    def _on_department_select(self, dept_name):
        self._set_field_text('department', dept_name)
        if hasattr(self.registration_screen.ids, 'section'):
            self.registration_screen.ids.section.text = ""

    def open_section_menu(self):
        self._dismiss_active_menu()
        dept = self.registration_screen.ids.department.text if hasattr(self.registration_screen.ids, 'department') else ""
        if dept == "Programs":
            sectors = ["MEAL", "PROTECTION", "EDUCATION", "LIVELIHOOD", "WASH", "NUTRITION", "GBV", "COMMUNICATION", "INFORMATION MANAGEMENT"]
        elif dept == "HR & Operation":
            sectors = ["HR", "LOGISTIC", "MAINTENANCE", "SECURITY"]
        elif dept == "Finance":
            sectors = ["Finance"]
        else:
            sectors = ["General"]

        items = [
            {"text": sec, "viewclass": "OneLineListItem", "on_release": lambda x=sec: self._set_field_text('section', x)}
            for sec in sectors
        ]
        self.active_menu = MDDropdownMenu(caller=self.registration_screen.ids.section, items=items, width_mult=4)
        self.active_menu.open()

    def open_employment_type_menu(self):
        self._dismiss_active_menu()
        types = ["Program Manager", "Manager", "Coordinator", "Officer", "Assistant", "Case Worker", "Intern", "Volunteer"]
        items = [
            {"text": emp, "viewclass": "OneLineListItem", "on_release": lambda x=emp: self._set_field_text('employment_type', x)}
            for emp in types
        ]
        self.active_menu = MDDropdownMenu(caller=self.registration_screen.ids.employment_type, items=items, width_mult=4)
        self.active_menu.open()

    def open_state_origin_menu(self):
        self._dismiss_active_menu()
        states = [
            "Abia State", "Adamawa State", "Akwa Ibom State", "Anambra State", "Bauchi State",
            "Bayelsa State", "Benue State", "Borno State", "Cross River State", "Delta State",
            "Ebonyi State", "Edo State", "Ekiti State", "Enugu State", "Gombe State",
            "Imo State", "Jigawa State", "Kaduna State", "Kano State", "Katsina State",
            "Kebbi State", "Kogi State", "Kwara State", "Lagos State", "Nasarawa State",
            "Niger State", "Ogun State", "Ondo State", "Osun State", "Oyo State",
            "Plateau State", "Rivers State", "Sokoto State", "Taraba State", "Yobe State", "Zamfara State"
        ]
        items = [
            {"text": st, "viewclass": "OneLineListItem", "on_release": lambda x=st: self._set_field_text('state_origin', x)}
            for st in states
        ]
        self.active_menu = MDDropdownMenu(caller=self.registration_screen.ids.state_origin, items=items, width_mult=4)
        self.active_menu.open()

    def open_marital_status_menu(self):
        self._dismiss_active_menu()
        statuses = ["Single", "Married", "Widow"]
        items = [
            {"text": status, "viewclass": "OneLineListItem", "on_release": lambda x=status: self._set_field_text('marital_status', x)}
            for status in statuses
        ]
        self.active_menu = MDDropdownMenu(caller=self.registration_screen.ids.marital_status, items=items, width_mult=4)
        self.active_menu.open()

    def open_blood_group_menu(self):
        self._dismiss_active_menu()
        groups = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
        items = [
            {"text": g, "viewclass": "OneLineListItem", "on_release": lambda x=g: self._set_field_text('blood_group', x)}
            for g in groups
        ]
        self.active_menu = MDDropdownMenu(caller=self.registration_screen.ids.blood_group, items=items, width_mult=4)
        self.active_menu.open()

    def open_date_picker(self):
        self._dismiss_active_menu()
        date_dialog = MDDatePicker()
        date_dialog.bind(on_save=self._on_date_save)
        date_dialog.open()

    def _on_date_save(self, instance, value, date_range):
        if hasattr(self.registration_screen.ids, 'dob'):
            self.registration_screen.ids.dob.text = str(value)

    def _set_field_text(self, field_id, text):
        if hasattr(self.registration_screen.ids, field_id):
            self.registration_screen.ids[field_id].text = text
        self._dismiss_active_menu()

    # -----------------------------
    # GPS Location
    # -----------------------------
    def gps_status(self, stype, status):
        pass

    # -----------------------------
    # Android Gallery / Image Picker
    # -----------------------------
    def _open_image_gallery(self, target):
        """Open the Android image picker without binding directly to the
        Activity object. Plyer's file chooser is used first because it survives
        Android activity recreation much more reliably than a raw
        startActivityForResult callback in packaged Kivy apps."""
        logger.info("Opening phone image picker for %s...", target)
        self._set_gallery_status(target, "")
        if platform != "android":
            self._set_gallery_status(target, "Image picker is available on Android.")
            return
        self._gallery_target = target
        try:
            # On a lot of devices the *first* tap only triggers Android's
            # storage-permission prompt in the background (no chooser
            # visibly opens), so the picker only actually shows up on the
            # second tap once permission is already granted. Request the
            # permission up front and open the chooser as soon as we get an
            # answer (granted or not - Storage Access Framework picking
            # generally still works either way), so one tap is enough.
            from android.permissions import request_permissions, Permission, check_permission
            needed = [Permission.READ_EXTERNAL_STORAGE]
            media_images = getattr(Permission, "READ_MEDIA_IMAGES", None)
            if media_images:
                needed.append(media_images)
            if all(check_permission(p) for p in needed):
                self._launch_filechooser()
            else:
                def _on_permission_result(permissions, grants):
                    Clock.schedule_once(lambda dt: self._launch_filechooser(), 0)
                request_permissions(needed, _on_permission_result)
        except Exception:
            # android.permissions unavailable - fall back to opening the
            # chooser directly, same as before.
            self._launch_filechooser()

    def _launch_filechooser(self):
        target = getattr(self, '_gallery_target', 'photo')
        try:
            from plyer import filechooser
            filechooser.open_file(
                on_selection=self._on_filechooser_selection,
                filters=["*.png", "*.jpg", "*.jpeg", "*.webp"],
                multiple=False,
            )
        except Exception as exc:
            logger.exception("Could not open image picker:")
            self._set_gallery_status(target, f"Could not open image picker: {exc}")

    def _on_filechooser_selection(self, selection):
        """Fallback callback used when the Android activity dispatcher does
        not expose ``bind``. Plyer returns a list of selected local paths.
        This is used for both registration photos and Timesheet signatures."""
        target = getattr(self, '_gallery_target', 'photo')
        try:
            if not selection:
                self._set_gallery_status(target, "No image was selected.")
                return
            source = selection[0] if isinstance(selection, (list, tuple)) else selection
            if not source:
                self._set_gallery_status(target, "No image was selected.")
                return
            source = str(source)
            if source.startswith('file://'):
                source = source[7:]
            if not os.path.isfile(source):
                raise RuntimeError("The selected image could not be accessed.")

            folder = 'staff_photos' if target == 'photo' else 'staff_signatures'
            prefix = 'photo' if target == 'photo' else 'signature'
            folder_path = os.path.join(APP_DIR, folder)
            os.makedirs(folder_path, exist_ok=True)
            destination = os.path.abspath(os.path.join(
                folder_path, f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            ))

            # Normalize the selected image to JPEG so the preview and Excel
            # embedding work consistently regardless of the original format.
            from PIL import Image
            with Image.open(source) as image:
                image = image.convert("RGB")
                image.thumbnail((1200, 1200), Image.Resampling.LANCZOS)
                image.save(destination, "JPEG", quality=88, optimize=True)

            if target == 'photo':
                self.photo_taken(destination)
            else:
                self.signature_captured(destination)
            self._set_gallery_status(target, "")
            logger.info("File chooser image imported successfully: %s", destination)
        except Exception as e:
            logger.exception("Failed to import image from file chooser:")
            self._set_gallery_status(target, f"Could not import image: {e}")

    def _set_gallery_status(self, target, message):
        """Shows a gallery import error/status message on the right screen,
        instead of failing silently with nothing visible on screen."""
        try:
            if target == 'photo' and hasattr(self, 'registration_screen') \
                    and hasattr(self.registration_screen.ids, 'photo_status_label'):
                self.registration_screen.ids.photo_status_label.text = message
            elif target == 'signature' and hasattr(self, 'timesheet_screen') \
                    and hasattr(self.timesheet_screen.ids, 'timesheet_status_label'):
                self.timesheet_screen.ids.timesheet_status_label.text = message
        except Exception:
            logger.exception("Failed to update gallery status label:")

    def _on_gallery_result(self, request_code, result_code, intent):
        if request_code != getattr(self, '_gallery_request_code', None):
            return
        target = getattr(self, '_gallery_target', 'photo')
        try:
            from jnius import autoclass
            if result_code != -1 or intent is None:
                logger.info("Gallery selection cancelled.")
                return
            uri = intent.getData()
            if uri is None:
                logger.warning("Gallery returned no image URI.")
                self._set_gallery_status(target, "No image was selected.")
                return
            folder = 'staff_photos' if target == 'photo' else 'staff_signatures'
            prefix = 'photo' if target == 'photo' else 'signature'
            folder_path = os.path.join(APP_DIR, folder)
            os.makedirs(folder_path, exist_ok=True)
            destination = os.path.abspath(os.path.join(
                folder_path, f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            ))

            activity = autoclass('org.kivy.android.PythonActivity').mActivity
            resolver = activity.getContentResolver()
            input_stream = resolver.openInputStream(uri)
            if input_stream is None:
                raise RuntimeError("Unable to read the selected gallery image.")

            # Decode via Android's own Bitmap decoder and re-encode to JPEG
            # bytes, rather than manually copying the InputStream into a
            # Python bytearray - pyjnius does not reliably copy Java's writes
            # into a Python-side bytearray passed as a byte[] argument, which
            # was silently producing empty/corrupt image files.
            BitmapFactory = autoclass('android.graphics.BitmapFactory')
            ByteArrayOutputStream = autoclass('java.io.ByteArrayOutputStream')
            CompressFormat = autoclass('android.graphics.Bitmap$CompressFormat')
            bitmap = BitmapFactory.decodeStream(input_stream)
            input_stream.close()
            if bitmap is None:
                raise RuntimeError("The selected file could not be read as an image.")
            # Downscale large phone photos/signatures before storing them. This
            # keeps the app responsive and prevents huge Excel exports.
            max_side = 1200
            bw, bh = int(bitmap.getWidth()), int(bitmap.getHeight())
            if max(bw, bh) > max_side:
                scale = max_side / float(max(bw, bh))
                Bitmap = autoclass("android.graphics.Bitmap")
                scaled = Bitmap.createScaledBitmap(
                    bitmap, max(1, int(bw * scale)), max(1, int(bh * scale)), True
                )
                if scaled is not None:
                    bitmap = scaled
            baos = ByteArrayOutputStream()
            bitmap.compress(CompressFormat.JPEG, 88, baos)
            java_bytes = baos.toByteArray()
            with open(destination, 'wb') as output_file:
                output_file.write(bytes(java_bytes))

            if target == 'photo':
                self.photo_taken(destination)
            else:
                self.signature_captured(destination)
            self._set_gallery_status(target, "")
            logger.info("Gallery image imported successfully: %s", destination)
        except Exception as e:
            logger.exception("Failed to import image from phone gallery:")
            self._set_gallery_status(target, f"Could not import image: {e}")
        finally:
            try:
                android_activity = getattr(self, '_android_activity', None)
                if android_activity is not None:
                    android_activity.unbind(on_activity_result=self._on_gallery_result)
            except Exception:
                pass

    # -----------------------------
    # Staff Registration Photo
    # -----------------------------
    def capture_photo(self):
        """Select the staff profile picture from the phone gallery (bound to
        the "Capture Photo" camera-icon button on the registration form)."""
        self._open_image_gallery('photo')

    def toggle_password_field(self, field_id, icon_button):
        """Shows/hides the text in a password field and flips its eye icon.
        Bound to the eye-icon buttons next to Create Password / Confirm
        Password on the registration form."""
        try:
            field = getattr(self.registration_screen.ids, field_id)
            field.password = not field.password
            icon_button.icon = "eye-off" if field.password else "eye"
        except Exception:
            logger.exception("Failed to toggle password visibility for %s:", field_id)

    def photo_taken(self, path):
        if path and os.path.exists(path):
            self.photo_path = path
            logger.info(f"Photo imported successfully to path: {path}")
            try:
                if hasattr(self.registration_screen.ids, 'photo_preview'):
                    img = self.registration_screen.ids.photo_preview
                    img.source = path
                    img.reload()
            except Exception:
                logger.exception("Failed to refresh photo preview UI:")

    # -----------------------------
    # Registration Submission
    # -----------------------------
    def _ensure_unique_id(self, user):
        """Some staff accounts were registered before auto-generated Unique IDs
        existed, so their unique_id column (index 29) is empty. This backfills
        one at login time and persists it, so their Dashboard and Attendance
        Report correctly show a Unique ID from now on instead of '-'."""
        try:
            if user and (len(user) <= 29 or not user[29]):
                new_id = self._generate_unique_staff_id()
                db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                try:
                    cursor.execute("UPDATE staff SET unique_id = ? WHERE id = ?", (new_id, user[0]))
                    conn.commit()
                except sqlite3.IntegrityError:
                    # Extremely unlikely collision - try once more with a fresh id.
                    new_id = self._generate_unique_staff_id()
                    cursor.execute("UPDATE staff SET unique_id = ? WHERE id = ?", (new_id, user[0]))
                    conn.commit()
                cursor.execute("SELECT * FROM staff WHERE id = ?", (user[0],))
                refreshed = cursor.fetchone()
                conn.close()
                logger.info(f"Backfilled Unique ID {new_id} for existing staff account (id={user[0]}).")
                return refreshed or user
        except Exception:
            logger.exception("Could not backfill unique_id for existing account:")
        return user

    @staticmethod
    def _generate_unique_staff_id():
        """Generates a short, human-readable unique staff ID, e.g. 'ROHI-260804-7F3K'.
        Collisions are astronomically unlikely, but insert_staff() will raise
        sqlite3.IntegrityError on the UNIQUE constraint if one ever happens,
        and submit_staff() retries generation in that case."""
        date_part = datetime.now().strftime("%y%m%d")
        code_part = "".join(random.choices("ABCDEFGHJKLMNPQRSTUVWXYZ23456789", k=4))
        return f"ROHI-{date_part}-{code_part}"

    def submit_staff(self):
        logger.info("Submitting staff registration form...")
        try:
            ids = self.registration_screen.ids

            def get_text(field_name):
                if hasattr(ids, field_name):
                    return getattr(ids, field_name).text
                return ""

            edit_mode = self.current_user is not None
            password = get_text("password").strip()
            confirm_password = get_text("confirm_password").strip()

            if password and confirm_password and password != confirm_password:
                logger.warning("Registration failed: Passwords do not match.")
                self._show_registration_error(
                    "Passwords Do Not Match",
                    "The password and confirm password fields do not match.\n\n"
                    "Please re-enter them."
                )
                return

            # Compulsory fields (marked with * on the form) must be filled
            # before the record is saved.
            missing_fields = [
                label for field_id, label in REQUIRED_REGISTRATION_FIELDS
                if not get_text(field_id).strip()
            ]
            if not edit_mode:
                if not password:
                    missing_fields.append("Create Password")
                if not confirm_password:
                    missing_fields.append("Confirm Password")

            if missing_fields:
                logger.warning("Registration rejected: missing compulsory fields: %s", ", ".join(missing_fields))
                self._show_registration_error(
                    "Missing Compulsory Fields",
                    "Please fill in the following compulsory fields marked with *:\n\n"
                    + "\n".join(f"\u2022 {f}" for f in missing_fields)
                )
                return

            if edit_mode and not password:
                password = self.current_user[28] if len(self.current_user) > 28 else ""

            email = get_text("email").strip().lower()
            staff_number = get_text("staff_number").strip().lower()

            # Check uniqueness BEFORE opening the INSERT/UPDATE transaction.
            # This prevents the old code from catching a UNIQUE email error and
            # incorrectly retrying the entire INSERT, which caused the second
            # error: "database is locked".
            current_id = self.current_user[0] if edit_mode else None
            if email_exists(email, exclude_id=current_id):
                logger.warning("Registration rejected: email already exists: %s", email)
                self._show_registration_error(
                    "Email Already Registered",
                    f"The email address '{email}' is already registered.\n\n"
                    "Please use a different email address."
                )
                return

            if staff_number and staff_number_exists(staff_number, exclude_id=current_id):
                logger.warning("Registration rejected: staff number already exists: %s", staff_number)
                self._show_registration_error(
                    "Staff Number Already Registered",
                    f"The staff number '{staff_number}' is already registered.\n\n"
                    "Please check the staff number and try again."
                )
                return

            staff_data = {
                "fullname": get_text("fullname"),
                "sex": get_text("sex"),
                "dob": get_text("dob"),
                "blood_group": get_text("blood_group"),
                "marital_status": get_text("marital_status"),
                "nationality": get_text("nationality"),
                "state_origin": get_text("state_origin"),
                "lga": get_text("lga"),
                "address": get_text("address"),
                "next_of_kin": get_text("next_of_kin"),
                "next_of_kin_phone": get_text("next_of_kin_phone"),
                "employment_type": get_text("employment_type"),
                "state_office": get_text("state_office"),
                "cluster": get_text("cluster"),
                "department": get_text("department"),
                "section": get_text("section"),
                "position": get_text("position"),
                "staff_number": staff_number,
                "phone": get_text("phone"),
                "email": email,
                "facebook": get_text("facebook"),
                "twitter": get_text("twitter"),
                "instagram": get_text("instagram"),
                "telegram": get_text("telegram"),
                "linkedin": get_text("linkedin"),
                "gps_coordinate": (getattr(self, "_selected_office_gps_coordinate", "")
                                    or get_text("gps_coordinate") or self.current_location),
                "photo": getattr(self, "photo_path", ""),
                "password": password,
                "genotype": get_text("genotype"),
                "reintegration_status": get_text("reintegration_status"),
            }

            if edit_mode:
                staff_id = self.current_user[0]
                update_staff(staff_id, staff_data)
                logger.info(
                    "Staff '%s' (id=%s) updated in SQLite database.",
                    staff_data["fullname"], staff_id
                )

                self.current_user = get_staff_by_id(staff_id)
                self._save_login_session(str(self.current_user[20]) if len(self.current_user) > 20 else email)
                self._submit_staff_registration_immediately(staff_data)
                email_val = self._populate_dashboard_from_current_user()
                if email_val:
                    self.verify_existing_checkin(email_val)
                self.update_dashboard_metrics()

                def redirect_to_dashboard(inst):
                    dialog.dismiss()
                    self.root.current = "dashboard"

                dialog = MDDialog(
                    title="Profile Updated",
                    text="Your staff profile was updated successfully.",
                    buttons=[MDFlatButton(text="OK", on_release=redirect_to_dashboard)]
                )
                dialog.open()

            else:
                new_unique_id = self._generate_unique_staff_id()
                staff_data["unique_id"] = new_unique_id

                # Only one staff registration is kept per phone. open_registration()
                # already blocks starting a new registration when one exists, but
                # this is a defensive safeguard so the local database can never end
                # up holding more than one registration record.
                try:
                    if get_staff_count() > 0:
                        clear_all_staff()
                except Exception:
                    logger.exception("Failed to clear previous registration before saving new one.")

                try:
                    insert_staff(staff_data)
                except sqlite3.IntegrityError as exc:
                    # Do NOT retry every IntegrityError. A UNIQUE email/staff
                    # number collision is a permanent validation error, not a
                    # transient lock. Only retry a generated unique_id collision.
                    message = str(exc).lower()
                    if "unique_id" not in message:
                        if "email" in message:
                            title = "Email Already Registered"
                            text = f"The email address '{email}' is already registered."
                        elif "staff_number" in message:
                            title = "Staff Number Already Registered"
                            text = f"The staff number '{staff_number}' is already registered."
                        else:
                            title = "Registration Error"
                            text = "A record with the same unique information already exists."
                        logger.warning("Registration rejected by SQLite: %s", exc)
                        self._show_registration_error(title, text)
                        return

                    staff_data["unique_id"] = self._generate_unique_staff_id()
                    insert_staff(staff_data)
                    new_unique_id = staff_data["unique_id"]

                logger.info(
                    "Staff '%s' inserted into SQLite database with Unique ID %s.",
                    staff_data["fullname"], new_unique_id
                )
                self._submit_staff_registration_immediately(staff_data)

                login_screen = self.root.get_screen("login")
                if hasattr(login_screen.ids, "email"):
                    login_screen.ids.email.text = staff_data["email"]
                if hasattr(login_screen.ids, "password"):
                    login_screen.ids.password.text = staff_data["password"]

                def redirect_to_login(inst):
                    dialog.dismiss()
                    self.root.current = "login"

                dialog = MDDialog(
                    title="Registration Successful",
                    text=(
                        "Your account was created successfully!\n\n"
                        f"Your Unique ID is: {new_unique_id}\n"
                        "(You'll also see this on your Dashboard.)\n\n"
                        "Please log in to continue."
                    ),
                    buttons=[MDFlatButton(text="OK", on_release=redirect_to_login)]
                )
                dialog.open()

        except sqlite3.OperationalError as exc:
            if "locked" in str(exc).lower():
                logger.exception("SQLite database remained locked during staff registration.")
                self._show_registration_error(
                    "Database Busy",
                    "The local database is busy. Please wait a few seconds and submit again."
                )
            else:
                logger.exception("Database error during staff registration.")
                self._show_registration_error(
                    "Database Error",
                    "The registration could not be saved. Please try again."
                )
        except Exception:
            logger.exception("Error during staff registration submission:")
            self._show_registration_error(
                "Registration Error",
                "An unexpected error occurred while saving the registration."
            )

    def _show_registration_error(self, title, text):
        """Display a registration error without crashing the app."""
        try:
            dialog = MDDialog(
                title=title,
                text=text,
                buttons=[MDFlatButton(text="OK", on_release=lambda *_: dialog.dismiss())]
            )
            dialog.open()
        except Exception:
            logger.exception("Unable to display registration error dialog.")


    # -----------------------------
    # Attendance Reports (period-filtered + all-time export)
    # -----------------------------
    def open_reports(self):
        self.root.current = "reports"
        if not self.current_user:
            return
        now = datetime.now()
        ids = self.reports_screen.ids
        ids.report_month_spinner.text = now.strftime("%B")
        ids.report_year_spinner.text = str(now.year)
        ids.report_day_spinner.text = "All Days (Monthly)"
        self.generate_report()

    def _populate_report_profile_header(self):
        if not self.current_user:
            return
        ids = self.reports_screen.ids
        user = self.current_user
        fullname = str(user[1]) if len(user) > 1 and user[1] else "-"
        employment_type = str(user[12]) if len(user) > 12 and user[12] else "-"
        cluster = str(user[14]) if len(user) > 14 and user[14] else "-"
        position = str(user[17]) if len(user) > 17 and user[17] else "-"
        base_gps = str(user[26]) if len(user) > 26 and user[26] else self.static_gps

        if hasattr(ids, 'report_name'):
            ids.report_name.text = f"Name: {fullname}"
        if hasattr(ids, 'report_employment_type'):
            ids.report_employment_type.text = f"Employment Type: {employment_type}"
        if hasattr(ids, 'report_cluster'):
            ids.report_cluster.text = f"Cluster: {cluster}"
        if hasattr(ids, 'report_position'):
            ids.report_position.text = f"Position: {position}"
        if hasattr(ids, 'report_base_gps'):
            ids.report_base_gps.text = f"Base GPS / Office Coordinate: {base_gps}"

    def _fetch_attendance_records(self, email_val, year=None, month=None, day=None):
        """Fetches attendance rows for a staff email. Passing year=None returns
        the FULL history to date (used by the all-time export)."""
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        if not email_val or not os.path.exists(db_path):
            return []
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        try:
            if year is None:
                cursor.execute(
                    "SELECT check_in_time, check_out_time, late_duration, attendance_status, "
                    "gps_location, check_out_gps_location "
                    "FROM attendance WHERE email = ? ORDER BY check_in_time ASC",
                    (email_val,)
                )
            else:
                if day and day != "All Days (Monthly)":
                    prefix = f"{year:04d}-{month:02d}-{int(day):02d}"
                else:
                    prefix = f"{year:04d}-{month:02d}-"
                cursor.execute(
                    "SELECT check_in_time, check_out_time, late_duration, attendance_status, "
                    "gps_location, check_out_gps_location "
                    "FROM attendance WHERE email = ? AND check_in_time LIKE ? ORDER BY check_in_time ASC",
                    (email_val, f"{prefix}%")
                )
            return cursor.fetchall()
        finally:
            conn.close()

    def generate_report(self):
        """Builds the on-screen period report (Present/Absent/Late + daily table)
        from real attendance records for the selected month/year/day."""
        try:
            if not self.current_user:
                return
            ids = self.reports_screen.ids
            email_val = str(self.current_user[20]) if len(self.current_user) > 20 else ""

            self._populate_report_profile_header()

            month_name = ids.report_month_spinner.text
            year = int(ids.report_year_spinner.text)
            day_sel = ids.report_day_spinner.text
            month = list(calendar.month_name).index(month_name)

            rows = self._fetch_attendance_records(email_val, year, month, day_sel)
            self._last_report_rows = rows

            present_days = len(set(r[0][:10] for r in rows if r[0]))
            late_count = sum(1 for r in rows if r[3] and "Late" in r[3])

            # Missing attendance rows are not automatic absences.
            absent_days = len(set(r[0][:10] for r in rows if r[0] and r[3] and "absent" in str(r[3]).lower()))

            if hasattr(ids, 'report_present_label'):
                ids.report_present_label.text = f"Present: {present_days} Day(s)"
            if hasattr(ids, 'report_absent_label'):
                ids.report_absent_label.text = f"Absent: {absent_days} Day(s)"
            if hasattr(ids, 'report_late_label'):
                ids.report_late_label.text = f"Total Late Occurrences: {late_count}"

            container = ids.report_rows_container
            container.clear_widgets()
            for check_in, check_out, late, status, gps, checkout_gps in rows:
                date_str = check_in[:10] if check_in else "-"
                in_time = check_in[11:16] if check_in and len(check_in) > 15 else "-"
                out_time = check_out[11:16] if check_out and len(check_out) > 15 else "Pending"
                in_within, _ = self._check_geofence(gps)
                in_status = "In:Within" if in_within else "In:Not Within"
                if checkout_gps:
                    out_within, _ = self._check_geofence(checkout_gps)
                    out_status = "Out:Within" if out_within else "Out:Not Within"
                else:
                    out_status = "Out:Pending"
                gps_status = f"{in_status} | {out_status}"

                row = MDBoxLayout(orientation='horizontal', spacing=dp(4), size_hint_y=None, height=dp(28))
                row.add_widget(MDLabel(text=date_str, font_style="Caption"))
                row.add_widget(MDLabel(text=f"{in_time} / {out_time}", font_style="Caption"))
                row.add_widget(MDLabel(text=late or "-", font_style="Caption"))
                row.add_widget(MDLabel(text=gps_status, font_style="Caption"))
                container.add_widget(row)

            period_desc = f"{month_name} {year}" if day_sel == "All Days (Monthly)" else f"{day_sel} {month_name} {year}"
            if hasattr(ids, 'report_status_label'):
                ids.report_status_label.text = f"{len(rows)} record(s) found for {period_desc}"
        except Exception:
            logger.exception("Error generating report:")

    def _staff_registration_headers_and_rows(self):
        """Build the complete staff registration export from every local staff
        record. Passwords and sync flags are never exported."""
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        conn = sqlite3.connect(db_path, timeout=10)
        try:
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(staff)")
            columns = [row[1] for row in cursor.fetchall()]
            if not columns:
                return [], []

            cursor.execute("SELECT * FROM staff ORDER BY id ASC")
            rows = cursor.fetchall()
            excluded = {"password", "synced"}
            display_names = {
                "id": "Database ID", "fullname": "Full Name", "sex": "Sex",
                "dob": "Date of Birth", "blood_group": "Blood Group",
                "marital_status": "Marital Status", "nationality": "Nationality",
                "state_origin": "State of Origin", "lga": "LGA",
                "address": "Residential Address", "next_of_kin": "Next of Kin",
                "next_of_kin_phone": "Next of Kin Phone", "employment_type": "Employment Type",
                "state_office": "State Office", "cluster": "Cluster", "department": "Department",
                "section": "Section", "position": "Position", "staff_number": "Staff Number",
                "phone": "Phone", "email": "Office Email", "facebook": "Facebook",
                "twitter": "Twitter", "instagram": "Instagram", "telegram": "Telegram",
                "linkedin": "LinkedIn", "gps_coordinate": "Office GPS Coordinate",
                "photo": "Staff Photo Path", "unique_id": "Unique ID",
                "genotype": "Genotype", "reintegration_status": "Reintegration Status",
            }
            included = [(idx, col) for idx, col in enumerate(columns) if col not in excluded]
            headers = [display_names.get(col, col.replace("_", " ").title()) for _, col in included]
            data_rows = []
            for row in rows:
                data_rows.append([
                    "-" if row[idx] in (None, "") else str(row[idx])
                    for idx, _ in included
                ])
            return headers, data_rows
        finally:
            conn.close()

    def _report_type_export(self, report_type, fmt):
        """Generates the selected report type (Staff Registration / Attendance
        - Selected Period / Attendance - Full History) in the given format and
        returns the saved filepath, or None on failure (status label is set
        either way)."""
        if not self.current_user:
            self._show_report_status("Please log in first.")
            return None
        try:
            reports_dir = self._get_download_dir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fullname_safe = str(self.current_user[1]).replace(" ", "_") if len(self.current_user) > 1 else "staff"

            if report_type == "staff_registration":
                headers, data_rows = self._staff_registration_headers_and_rows()
                title = "ROHI Staff Registration Details"
                scope = "staff-registration"
            elif report_type == "attendance_history":
                email_val = str(self.current_user[20]) if len(self.current_user) > 20 else ""
                rows = self._fetch_attendance_records(email_val)
                headers, data_rows = self._report_headers_and_rows(rows)
                title = "ROHI Attendance Report (All Time To Date)"
                scope = "all-time-to-date"
            else:  # "attendance" - currently selected report period
                headers, data_rows = self._report_headers_and_rows(getattr(self, '_last_report_rows', []))
                title = "ROHI Attendance Report (Selected Period)"
                scope = "selected-period"

            ext = "pdf" if fmt == "pdf" else "xlsx"
            filename = f"{scope}_{fullname_safe}_{timestamp}.{ext}"
            filepath = os.path.join(reports_dir, filename)

            if fmt == "pdf":
                self._build_pdf(filepath, title, headers, data_rows, password=REPORT_EXPORT_PASSWORD)
            else:
                if report_type == "staff_registration":
                    filepath = self._export_staff_template()
                elif report_type in ("attendance", "attendance_history"):
                    filepath = self._export_attendance_template(
                        getattr(self, "_last_report_rows", []) if report_type == "attendance"
                        else self._fetch_attendance_records(str(self.current_user[20]))
                    )
                else:
                    self._build_excel(filepath, headers, data_rows, sheet_title=title[:31],
                                       password=REPORT_EXPORT_PASSWORD)

            self._last_export_path = filepath
            self._show_report_status(f"Exported: {filepath}")
            if fmt == "xlsx":
                sync_type = "staff" if report_type == "staff_registration" else "attendance"
                self._queue_excel_auto_sync(filepath, sync_type)
            return filepath
        except ImportError as e:
            logger.exception("Missing export library:")
            self._show_report_status(f"Export library not installed ({e}). Add it to buildozer.spec requirements.")
            return None
        except Exception:
            logger.exception("Error exporting report:")
            self._show_report_status("Error exporting report - see logs.")
            return None

    REPORT_TYPE_OPTIONS = [
        ("Attendance - Selected Period", "attendance"),
        ("Attendance - Full History", "attendance_history"),
        ("Complete Staff Registration", "staff_registration"),
    ]

    def open_report_type_menu(self):
        items = [
            {
                "text": label,
                "viewclass": "OneLineListItem",
                "on_release": lambda x=label: self._select_report_type(x),
            }
            for label, _ in self.REPORT_TYPE_OPTIONS
        ]
        self._dismiss_active_menu()
        self.active_menu = MDDropdownMenu(
            caller=self.reports_screen.ids.report_type_field,
            items=items,
            width_mult=5,
        )
        self.active_menu.open()

    def _select_report_type(self, label):
        self.reports_screen.ids.report_type_field.text = label
        self._dismiss_active_menu()

    def generate_selected_report(self):
        """One export button for the three report choices.

        All three choices generate an Excel workbook so the same button has a
        predictable result. The mail icon beside it can then share the exact
        file that was generated.
        """
        ids = self.reports_screen.ids
        label = ids.report_type_field.text.strip() if hasattr(ids, 'report_type_field') else ""
        type_map = dict(self.REPORT_TYPE_OPTIONS)
        report_type = type_map.get(label)
        if not report_type:
            self._show_report_status("Select a report type first.")
            return
        self._report_type_export(report_type, fmt="excel")

    def generate_and_email_report(self):
        """Generate the Attendance report and open Android's email/share composer."""
        self.generate_selected_report()
        return self.email_attendance_report()

    def _report_headers_and_rows(self, rows):
        """Converts raw attendance rows into the export table, matching ROHI's
        required attendance-report column layout exactly:
        SN, Unique Id, Date, Name, Sex, Position, State Office,
        State office Cordinates, Cluster, Check in, Late Hour, check in Time,
        Check in current, Meter Range, Check in Status, check out,
        overtime ours, check out time, check out current coordinated,
        check out gps current State, Verified Check in, Verified check out."""
        if not self.current_user:
            return [], []
        user = self.current_user
        fullname = str(user[1]) if len(user) > 1 and user[1] else "-"
        sex = str(user[2]) if len(user) > 2 and user[2] else "-"
        state_office = str(user[13]) if len(user) > 13 and user[13] else "-"
        cluster = str(user[14]) if len(user) > 14 and user[14] else "-"
        position = str(user[17]) if len(user) > 17 and user[17] else "-"
        # Unique Id: the auto-generated ID from registration (column 29).
        # Falls back to staff_number for any pre-existing records created
        # before unique-ID generation was added.
        unique_id = str(user[29]) if len(user) > 29 and user[29] else "-"
        if unique_id == "-":
            unique_id = str(user[18]) if len(user) > 18 and user[18] else "-"
        base_gps = str(user[26]) if len(user) > 26 and user[26] else self.static_gps

        staff_number = str(user[18]) if len(user) > 18 and user[18] else "-"

        headers = ["SN", "Unique Id", "Staff ID", "Date", "Name", "Sex", "Positìon", "State Office",
                   "State office Cordinates", "Cluster", "Check in", "Late Hour", "check in Time",
                   "Check in current", "Meter Range", "Check in Status ", "check out",
                   "overtime ours", "check out time", "check out current  coordinated ",
                   "check out gps current State", "Verified Check in", "Verified check out"]

        data_rows = []
        for sn, (check_in, check_out, late, status, gps, checkout_gps) in enumerate(rows, start=1):
            date_str = check_in[:10] if check_in else "-"
            in_time = check_in[11:19] if check_in and len(check_in) > 15 else "-"
            out_time = check_out[11:19] if check_out and len(check_out) > 15 else "-"

            check_in_word = "present" if check_in else "absent"
            check_out_word = "present" if check_out else ("pending" if check_in else "absent")

            in_within, in_distance = self._check_geofence(gps) if gps else (False, None)
            meter_range = f"{in_distance:.0f} m" if in_distance is not None else "-"
            in_gps_status = "Within office range" if in_within else "Not within office Range"
            verified_in = "Yes" if in_within else ("No" if gps else "-")

            if checkout_gps:
                out_within, out_distance = self._check_geofence(checkout_gps)
                out_state = state_office if out_within else "Unverified"
                verified_out = "Yes" if out_within else "No"
            else:
                out_within, out_distance = False, None
                out_state = "-"
                verified_out = "-"

            overtime_hours = "-"
            if check_in and check_out:
                try:
                    t_in = datetime.strptime(check_in, "%Y-%m-%d %H:%M:%S")
                    t_out = datetime.strptime(check_out, "%Y-%m-%d %H:%M:%S")
                    worked = (t_out - t_in).total_seconds() / 3600
                    overtime_hours = f"{max(worked - 8, 0):.1f}"
                except Exception:
                    overtime_hours = "-"

            data_rows.append([
                sn, unique_id, staff_number, date_str, fullname, sex, position, state_office,
                base_gps, cluster, check_in_word, late or "On Time", in_time,
                gps or "-", meter_range, in_gps_status, check_out_word,
                overtime_hours, out_time, checkout_gps or "-",
                out_state, verified_in, verified_out
            ])
        return headers, data_rows

    def _build_pdf(self, filepath, title, headers, data_rows, password=None):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(title, styles['Title']), Spacer(1, 12)]
        table = Table([headers] + [[str(c) for c in row] for row in data_rows], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#227A29')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
        ]))
        elements.append(table)
        doc.build(elements)

        if password:
            self._encrypt_pdf_open_password(filepath, password)

    @staticmethod
    def _encrypt_pdf_open_password(filepath, password):
        """Locks the exported PDF so it can't be opened without the password.
        Uses pypdf, a pure-Python library (no compiled/native dependency),
        so it's safe to bundle in the Android build."""
        try:
            from pypdf import PdfReader, PdfWriter

            reader = PdfReader(filepath)
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            writer.encrypt(password)
            with open(filepath, "wb") as f:
                writer.write(f)
            logger.info(f"Applied open-password encryption to {filepath}")
        except ImportError:
            logger.warning("pypdf not installed - exported PDF report is NOT password protected. Add 'pypdf' to buildozer.spec requirements.")
        except Exception:
            logger.exception(f"Could not apply open-password encryption to {filepath}")

    def _template_path(self, filename):
        return os.path.join(APP_DIR, "templates", filename)

    def _copy_template(self, filename, filepath):
        source = self._template_path(filename)
        if not os.path.exists(source):
            raise FileNotFoundError(f"Missing report template: {source}")
        shutil.copy2(source, filepath)

    def _export_staff_template(self, for_sync=False):
        """Export all staff using the uploaded ROHI Staff on IMS workbook as
        the visual template, preserving logo, borders, row heights and widths."""
        reports_dir = self._get_download_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if for_sync:
            filepath = os.path.join(reports_dir, "ROHI_Staff_on_IMS_AutoSync.xlsx")
        else:
            filepath = os.path.join(reports_dir, f"ROHI_Staff_on_IMS_{timestamp}.xlsx")
        self._copy_template("Rohi_Staff on IMS.xlsx", filepath)

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers, data_rows = self._staff_registration_headers_and_rows()
        # Template header is fixed at row 3. Use the template's exact order;
        # write values beginning at row 4 and keep its existing dimensions.
        template_headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
        index_map = {str(h).strip().lower(): i for i, h in enumerate(headers)}
        for r in range(4, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).value = None

        from copy import copy as _copy
        template_row_height = ws.row_dimensions[4].height
        template_styles = {
            c: _copy(ws.cell(4, c)._style) for c in range(1, ws.max_column + 1)
        }
        template_alignment = {
            c: _copy(ws.cell(4, c).alignment) for c in range(1, ws.max_column + 1)
        }
        template_number_formats = {
            c: ws.cell(4, c).number_format for c in range(1, ws.max_column + 1)
        }
        for r_idx, row in enumerate(data_rows, start=4):
            if r_idx > ws.max_row:
                ws.insert_rows(r_idx)
            if template_row_height is not None:
                ws.row_dimensions[r_idx].height = template_row_height
            for c_idx, h in enumerate(template_headers, start=1):
                cell = ws.cell(r_idx, c_idx)
                if r_idx > 4:
                    cell._style = _copy(template_styles.get(c_idx))
                    cell.alignment = _copy(template_alignment.get(c_idx))
                    cell.number_format = template_number_formats.get(c_idx, "General")
                pos = index_map.get(str(h).strip().lower())
                value = row[pos] if pos is not None and pos < len(row) else "-"
                cell.value = value

        wb.save(filepath)
        try:
            self._encrypt_xlsx_open_password(filepath, REPORT_EXPORT_PASSWORD)
        except Exception:
            logger.exception("Could not password-protect staff template export.")
        self._last_export_path = filepath
        return filepath

    def _export_attendance_template(self, rows, title_suffix="Selected"):
        """Export attendance using the uploaded attendance workbook template."""
        reports_dir = self._get_download_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(reports_dir, f"Rohi_Attendance_Report_{_safe_filename(title_suffix)}_{timestamp}.xlsx")
        self._copy_template("Rohi_Attendance Report.xlsx", filepath)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers, data_rows = self._report_headers_and_rows(rows)

        # Keep the template's header row (row 2) and widths. Data starts at row 3.
        for c, h in enumerate(headers, start=1):
            ws.cell(2, c).value = h
        # Clear old data rows and append new rows without altering the template
        # column widths or logo dimensions.
        if ws.max_row >= 3:
            for row_num in range(3, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row_num, c).value = None
        for r_idx, row in enumerate(data_rows, start=3):
            if r_idx > ws.max_row:
                ws.insert_rows(r_idx)
            for c_idx, value in enumerate(row, start=1):
                ws.cell(r_idx, c_idx).value = value
        wb.save(filepath)
        try:
            self._encrypt_xlsx_open_password(filepath, REPORT_EXPORT_PASSWORD)
        except Exception:
            logger.exception("Could not password-protect attendance template export.")
        self._last_export_path = filepath
        return filepath

    def _export_timesheet_template(self, for_sync=False):
        """Populate the exact uploaded Corrected Timesheet.xlsx template.
        Column widths, row heights, logo, thick/thin borders and merged cells
        remain from the supplied workbook."""
        rows = getattr(self, "_last_timesheet_rows", [])
        if not rows:
            return None
        ids = self.timesheet_screen.ids
        month_name = ids.timesheet_month_spinner.text
        year = ids.timesheet_year_spinner.text
        user = self.current_user
        fullname = str(user[1]) if user and len(user) > 1 and user[1] else "Staff"
        staff_id = str(user[18]) if user and len(user) > 18 and user[18] else "-"
        cluster = str(user[14]) if user and len(user) > 14 and user[14] else "-"
        state_office = str(user[13]) if user and len(user) > 13 and user[13] else "-"
        nationality = str(user[6]) if user and len(user) > 6 and user[6] else "-"

        reports_dir = self._get_download_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if for_sync:
            filename = f"ROHI_Timesheet_AutoSync_{_safe_filename(fullname)}_{month_name}_{year}.xlsx"
        else:
            filename = f"Timesheet_{_safe_filename(fullname)}_{month_name}_{year}_{timestamp}.xlsx"
        filepath = os.path.join(reports_dir, filename)
        self._copy_template("Corrected Timesheet.xlsx", filepath)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        ws["C3"] = f"{month_name.upper()} {year}"
        ws["B4"] = fullname
        ws["B5"] = staff_id
        ws["F3"] = cluster or state_office
        ws["F4"] = state_office
        ws["F5"] = nationality

        # Exact 31-day pay period rows from the uploaded template.
        for idx, row in enumerate(rows, start=8):
            date_label, day_name, hours_cell = row
            ws.cell(idx, 1).value = date_label
            ws.cell(idx, 2).value = day_name
            # Weekend rows are merged in the source template; only write the
            # value into the merged range's top-left cell.
            try:
                ws.cell(idx, 3).value = hours_cell
            except AttributeError:
                pass
        # The supplied template already keeps the lower cells of merged
        # weekend ranges empty. Do not write to MergedCell placeholders.

        ws["C39"] = "=SUM(C8:C38)"
        # Signature: keep template column widths and place a scaled image over
        # B46:D47 so it fits the existing signature area without distorting it.
        sig_path = getattr(self, "signature_path", None)
        if sig_path and os.path.exists(sig_path):
            try:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(sig_path)
                max_w, max_h = 190, 55
                ratio = min(max_w / img.width, max_h / img.height, 1.0)
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
                ws.add_image(img, "B46")
            except Exception:
                logger.exception("Could not embed signature image in timesheet template.")

        # Keep the supplied template's signature lines.
        ws["A46"] = "Employee Signature: ____________________________"
        ws["A49"] = "Supervisor Signature: ____________________________"
        wb.save(filepath)
        self._last_export_path = filepath
        return filepath

    def _build_excel(self, filepath, headers, data_rows, sheet_title="Report", password=None):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        # ---- Logo + title band (rows 1-2), then the exact template header
        # layout starts at row 3 so the column structure/colors below match
        # the ROHI attendance report template exactly. ----
        header_row = 1
        app_dir = os.path.dirname(os.path.abspath(__file__))
        logo_candidates = [
            os.path.join(app_dir, "rohi_logo.png"),
            os.path.join(app_dir, "Images", "rohi_logo.png"),
        ]
        logo_path = next((candidate for candidate in logo_candidates if os.path.exists(candidate)), None)
        if logo_path:
            try:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(logo_path)
                img.width, img.height = 32, 32
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = 38
            except Exception:
                logger.exception("Could not embed ROHI logo in exported report:")
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=2, value=f"RESTORATION OF HOPE INITIATIVE (ROHI) - {sheet_title}")
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 6  # thin spacer row
        header_row = 3

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=h)
        # Green fill + bold white text ONLY on columns D through T (Name
        # through "check out gps current State") - matching the template
        # exactly, where SN/Unique Id/Date and the two "Verified" columns
        # are left unstyled.
        for col_idx in range(4, len(headers) + 1):  # style report fields; ID columns remain plain
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="227A29", end_color="227A29", fill_type="solid")

        for row in data_rows:
            ws.append(row)

        # ---- Column widths matching the template ----
        template_widths = {
            "D": 14, "E": 13.05, "H": 16.95, "J": 9.95, "K": 16.01, "L": 11.97,
            "N": 18.56, "O": 11.97, "P": 27.98, "Q": 15.33, "R": 16.41,
            "T": 23.95, "U": 24.48, "V": 28, "W": 72.24,
        }
        for letter, width in template_widths.items():
            ws.column_dimensions[letter].width = width
        # Fall back to auto-fit for any column not explicitly sized above.
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            if letter in template_widths:
                continue
            values = [headers[col_idx - 1]] + [str(r[col_idx - 1]) for r in data_rows if r[col_idx - 1] is not None]
            max_len = max((len(v) for v in values), default=8)
            ws.column_dimensions[letter].width = max_len + 2

        if password:
            # Sheet-level lock (works even if the full-file encryption below
            # can't run) - prevents editing without the password.
            ws.protection.sheet = True
            ws.protection.password = password

        wb.save(filepath)

        if password:
            self._encrypt_xlsx_open_password(filepath, password)

    @staticmethod
    def _encrypt_xlsx_open_password(filepath, password):
        """Locks the .xlsx file itself so it cannot be OPENED at all without
        the password (not just edit-protection). Requires msoffcrypto-tool +
        cryptography, both listed in buildozer.spec. If either isn't
        available on this build, the file still has the sheet-level lock
        applied above, and the export is not blocked."""
        try:
            import io
            import msoffcrypto
            from msoffcrypto.format.ooxml import OOXMLFile

            with open(filepath, "rb") as f:
                plain_bytes = f.read()

            office_file = OOXMLFile(io.BytesIO(plain_bytes))
            encrypted = io.BytesIO()
            office_file.encrypt(password, encrypted)

            with open(filepath, "wb") as f:
                f.write(encrypted.getvalue())
            logger.info(f"Applied open-password encryption to {filepath}")
        except ImportError:
            logger.warning(
                "msoffcrypto-tool not installed - exported report only has "
                "sheet-edit protection, not a full open-password lock. "
                "Add 'msoffcrypto-tool' and 'cryptography' to buildozer.spec requirements."
            )
        except Exception:
            logger.exception(
                f"Could not apply open-password encryption to {filepath}; "
                "file was still saved with sheet-level protection only."
            )

    def _get_download_dir(self):
        """Resolve the phone's public Download folder so generated reports show up
        where the user actually looks for downloaded files, instead of being
        buried inside the app's private storage."""
        if platform == "android":
            try:
                from jnius import autoclass
                Environment = autoclass('android.os.Environment')
                download_dir = Environment.getExternalStoragePublicDirectory(
                    Environment.DIRECTORY_DOWNLOADS
                ).getAbsolutePath()
            except Exception:
                logger.exception("Falling back to hardcoded Android Download path:")
                download_dir = "/storage/emulated/0/Download"
        else:
            download_dir = os.path.join(os.path.expanduser("~"), "Downloads")

        os.makedirs(download_dir, exist_ok=True)
        return download_dir

    def _show_report_status(self, message):
        if hasattr(self, 'reports_screen') and hasattr(self.reports_screen.ids, 'report_status_label'):
            self.reports_screen.ids.report_status_label.text = message
        logger.info(message)

    def _export_attendance_rows(self, rows, fmt, scope):
        try:
            if not self.current_user:
                self._show_report_status("Please log in first.")
                return
            # Even with zero attendance rows (e.g. a brand-new staff member who
            # hasn't checked in yet), still produce a downloadable file with the
            # full header row so the report format is available from day one -
            # it just has no data rows underneath until check-ins happen.
            headers, data_rows = self._report_headers_and_rows(rows)
            reports_dir = self._get_download_dir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fullname_safe = (str(self.current_user[1]).replace(" ", "_")
                              if self.current_user and len(self.current_user) > 1 else "staff")

            if fmt == 'pdf':
                filename = f"Attendance_{scope}_{fullname_safe}_{timestamp}.pdf"
                filepath = os.path.join(reports_dir, filename)
                self._build_pdf(filepath, f"ROHI Attendance Report ({scope.replace('-', ' ').title()})",
                                 headers, data_rows, password=REPORT_EXPORT_PASSWORD)
            else:
                filename = f"Attendance_{scope}_{fullname_safe}_{timestamp}.xlsx"
                filepath = os.path.join(reports_dir, filename)
                self._build_excel(filepath, headers, data_rows, sheet_title="Attendance",
                                   password=REPORT_EXPORT_PASSWORD)

            self._last_export_path = filepath
            self._show_report_status(f"Exported (password protected): {filepath}")
        except ImportError as e:
            logger.exception("Missing export library:")
            self._show_report_status(f"Export library not installed ({e}). Add it to buildozer.spec requirements.")
        except Exception:
            logger.exception("Error exporting attendance report:")
            self._show_report_status("Error exporting report - see logs.")

    def export_report_pdf(self):
        self._export_attendance_rows(getattr(self, '_last_report_rows', []), fmt='pdf', scope='selected-period')

    def _share_file_via_android(self, filepath, subject, body, status_fn):
        """Share a generated file through Android's mail/share chooser.

        ACTION_SENDTO cannot carry file attachments. The previous code used
        ACTION_SENDTO and also called Intent.createChooser(), which produced the
        phone error about ``No static methods called createChooser``. This
        implementation uses ACTION_SEND plus an Android content URI, then opens
        the chooser as an ordinary ACTION_CHOOSER intent (no static Java method).
        """
        if not filepath or not os.path.exists(filepath):
            status_fn("Generate the report first, then use Email.")
            return False
        if platform != 'android':
            status_fn("Email sharing is available on Android only.")
            return False

        try:
            from jnius import autoclass
            Intent = autoclass('android.content.Intent')
            Uri = autoclass('android.net.Uri')
            BuildVersion = autoclass('android.os.Build$VERSION')
            File = autoclass('java.io.File')
            StrictMode = autoclass('android.os.StrictMode')

            filename = os.path.basename(filepath)
            lower = filename.lower()
            if lower.endswith('.pdf'):
                mime_type = 'application/pdf'
            elif lower.endswith('.xlsx'):
                mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            else:
                mime_type = 'application/octet-stream'

            activity = autoclass('org.kivy.android.PythonActivity').mActivity
            resolver = activity.getContentResolver()
            content_uri = None

            # Android 10+ / API 29+: put a shareable copy in MediaStore
            # Downloads and obtain a content:// URI that mail apps can read.
            if BuildVersion.SDK_INT >= 29:
                MediaStore = autoclass('android.provider.MediaStore')
                ContentValues = autoclass('android.content.ContentValues')
                values = ContentValues()
                values.put(MediaStore.MediaColumns.DISPLAY_NAME, filename)
                values.put(MediaStore.MediaColumns.MIME_TYPE, mime_type)
                values.put(MediaStore.MediaColumns.RELATIVE_PATH, 'Download/ROHI Attendance')
                values.put(MediaStore.MediaColumns.IS_PENDING, 1)
                collection = MediaStore.Downloads.EXTERNAL_CONTENT_URI
                content_uri = resolver.insert(collection, values)
                if content_uri is None:
                    raise RuntimeError('Android could not create a shareable Downloads URI.')
                output_stream = resolver.openOutputStream(content_uri, 'w')
                if output_stream is None:
                    raise RuntimeError('Android could not open the shareable file.')
                with open(filepath, 'rb') as source:
                    while True:
                        chunk = source.read(1024 * 1024)
                        if not chunk:
                            break
                        output_stream.write(chunk)
                output_stream.close()
                values_pending = ContentValues()
                values_pending.put(MediaStore.MediaColumns.IS_PENDING, 0)
                resolver.update(content_uri, values_pending, None, None)
            else:
                # Older Android: the generated file is already in public
                # Downloads. Use a MediaStore content URI where possible.
                MediaStoreFiles = autoclass('android.provider.MediaStore$Files')
                ContentValues = autoclass('android.content.ContentValues')
                values = ContentValues()
                values.put(MediaStore.MediaColumns.DATA, filepath)
                values.put(MediaStore.MediaColumns.DISPLAY_NAME, filename)
                values.put(MediaStore.MediaColumns.MIME_TYPE, mime_type)
                content_uri = resolver.insert(MediaStoreFiles.getContentUri('external'), values)
                if content_uri is None:
                    # Last-resort compatibility path for old Android releases.
                    try:
                        StrictMode.setVmPolicy(StrictMode.VmPolicy.Builder().build())
                    except Exception:
                        pass
                    content_uri = Uri.fromFile(File(filepath))

            send_intent = Intent(Intent.ACTION_SEND)
            send_intent.setType(mime_type)
            send_intent.putExtra(Intent.EXTRA_SUBJECT, subject)
            send_intent.putExtra(Intent.EXTRA_TEXT, body)
            send_intent.putExtra(Intent.EXTRA_STREAM, content_uri)
            send_intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)

            chooser = Intent(Intent.ACTION_CHOOSER)
            chooser.putExtra(Intent.EXTRA_INTENT, send_intent)
            chooser.putExtra(Intent.EXTRA_TITLE, 'Send Report via Email')
            chooser.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
            activity.startActivity(chooser)
            status_fn(f"Ready to send: {filename}")
            return True
        except Exception as e:
            logger.exception("Could not share generated file through Android:")
            status_fn(f"Could not open email app: {e}")
            return False

    def email_last_export(self, status_fn=None):
        """Open Android's share/email composer with the generated Attendance XLSX attached."""
        status_fn = status_fn or self._show_report_status
        filepath = getattr(self, '_last_export_path', None)
        if not filepath or not os.path.exists(filepath) or "attendance" not in os.path.basename(filepath).lower():
            self.generate_selected_report()
            filepath = getattr(self, '_last_export_path', None)
        return self._share_file_via_android(
            filepath,
            "ROHI Attendance Report",
            "Attached is the generated ROHI Attendance Report.",
            status_fn,
        )

    def email_attendance_report(self):
        return self.email_last_export(self._set_report_send_status)

    def email_timesheet_export(self):
        """Open Android's share/email composer with the generated Timesheet XLSX attached."""
        path = getattr(self, '_last_export_path', None)
        if not path or not os.path.exists(path) or 'timesheet' not in os.path.basename(path).lower():
            path = self.export_timesheet_excel()
        return self._share_file_via_android(
            path,
            "ROHI Timesheet",
            "Attached is the generated ROHI Timesheet.",
            self._set_timesheet_send_status,
        )

    def email_leave_report(self):
        """Open Android's share/email composer with the generated Leave XLSX attached."""
        path = getattr(self, '_last_export_path', None)
        if not path or not os.path.exists(path) or 'leave_report' not in os.path.basename(path).lower():
            path = self.export_leave_excel()
        return self._share_file_via_android(
            path,
            "ROHI Leave Management Report",
            "Attached is the generated ROHI Leave Management report.",
            self._set_leave_send_status,
        )

    def export_report_excel(self):
        self._export_attendance_rows(getattr(self, '_last_report_rows', []), fmt='excel', scope='selected-period')

    def export_all_attendance_pdf(self):
        email_val = str(self.current_user[20]) if self.current_user and len(self.current_user) > 20 else ""
        rows = self._fetch_attendance_records(email_val)
        self._export_attendance_rows(rows, fmt='pdf', scope='all-time-to-date')

    def export_all_attendance_excel(self):
        email_val = str(self.current_user[20]) if self.current_user and len(self.current_user) > 20 else ""
        rows = self._fetch_attendance_records(email_val)
        self._export_attendance_rows(rows, fmt='excel', scope='all-time-to-date')

    # -----------------------------
    # Leave Management
    # -----------------------------
    LEAVE_DEFAULT_BALANCES = {
        "Annual Leave": 20,
        "Sick Leave": 10,
        "Maternity Leave": 84,
        "Paternity Leave": 14,
        "Compassionate Leave": 5,
        "Other Leave": 5,
    }

    def open_leave(self):
        """Open Leave Management and refresh the current staff's balance/history."""
        if not self.current_user:
            self.root.current = "login"
            return
        self.root.current = "leave"
        ids = self.leave_screen.ids
        fullname = str(self.current_user[1]) if len(self.current_user) > 1 and self.current_user[1] else "-"
        position = str(self.current_user[17]) if len(self.current_user) > 17 and self.current_user[17] else "-"
        email = str(self.current_user[20]) if len(self.current_user) > 20 and self.current_user[20] else ""
        ids.leave_staff.text = f"Staff: {fullname}"
        ids.leave_position.text = f"Position: {position}"
        now = datetime.now()
        if not ids.leave_start.text:
            ids.leave_start.text = now.strftime("%Y-%m-%d")
        if not ids.leave_end.text:
            ids.leave_end.text = now.strftime("%Y-%m-%d")
        self._refresh_leave_summary(email)

    def _refresh_leave_summary(self, email=None):
        try:
            if not email and self.current_user and len(self.current_user) > 20:
                email = str(self.current_user[20])
            year = datetime.now().year
            used = {
                leave_type: get_leave_usage(email, leave_type, year)
                for leave_type in self.LEAVE_DEFAULT_BALANCES
            }
            ids = self.leave_screen.ids
            ids.annual_balance.text = f"{max(0, self.LEAVE_DEFAULT_BALANCES['Annual Leave'] - used['Annual Leave'])} days"
            ids.sick_balance.text = f"{max(0, self.LEAVE_DEFAULT_BALANCES['Sick Leave'] - used['Sick Leave'])} days"
            ids.maternity_balance.text = f"{max(0, self.LEAVE_DEFAULT_BALANCES['Maternity Leave'] - used['Maternity Leave'])} days"
            ids.paternity_balance.text = f"{max(0, self.LEAVE_DEFAULT_BALANCES['Paternity Leave'] - used['Paternity Leave'])} days"
            ids.compassionate_balance.text = f"{max(0, self.LEAVE_DEFAULT_BALANCES['Compassionate Leave'] - used['Compassionate Leave'])} days"
            ids.other_balance.text = f"{max(0, self.LEAVE_DEFAULT_BALANCES['Other Leave'] - used['Other Leave'])} days"
            rows = get_leave_requests(email)
            counts = get_leave_status_counts(email)
            ids.pending_count.text = str(counts.get("Pending", 0))
            ids.approved_count.text = str(counts.get("Approved", 0))
            ids.rejected_count.text = str(counts.get("Rejected", 0))
            try:
                dash_ids = self.dashboard_screen.ids
                dash_ids.leave_pending_count.text = str(counts.get("Pending", 0))
                dash_ids.leave_approved_count.text = str(counts.get("Approved", 0))
                dash_ids.leave_rejected_count.text = str(counts.get("Rejected", 0))
            except Exception:
                pass
            if not rows:
                ids.leave_history.text = "No leave requests yet."
                return
            lines = []
            for row in rows:
                # id, type, start, end, days, reason, status, manager_comment, submitted_at
                status = row[6] or "Pending"
                comment = f" — {row[7]}" if row[7] else ""
                lines.append(f"{row[1]} | {row[2]} to {row[3]} | {row[4]} day(s) | {status}{comment}")
            ids.leave_history.text = "\n".join(lines)
        except Exception:
            logger.exception("Unable to refresh leave summary")

    def submit_leave_request(self):
        """Validate and save a leave request as Pending."""
        if not self.current_user:
            return
        ids = self.leave_screen.ids
        leave_type = ids.leave_type_spinner.text.strip()
        start_text = ids.leave_start.text.strip()
        end_text = ids.leave_end.text.strip()
        reason = ids.leave_reason.text.strip()

        try:
            start = datetime.strptime(start_text, "%Y-%m-%d")
            end = datetime.strptime(end_text, "%Y-%m-%d")
        except ValueError:
            ids.leave_message.text = "Enter dates as YYYY-MM-DD."
            return

        if end < start:
            ids.leave_message.text = "End date cannot be before start date."
            return
        if start.date() < datetime.now().date():
            ids.leave_message.text = "Leave start date cannot be in the past."
            return
        days = (end.date() - start.date()).days + 1
        email = str(self.current_user[20]) if len(self.current_user) > 20 else ""
        fullname = str(self.current_user[1]) if len(self.current_user) > 1 else ""

        # Prevent submitting beyond the configured balance for leave types with a balance.
        limits = self.LEAVE_DEFAULT_BALANCES
        if leave_type in limits:
            used = get_leave_usage(email, leave_type, start.year)
            remaining = max(0, limits[leave_type] - used)
            if days > remaining:
                ids.leave_message.text = f"Insufficient {leave_type} balance. Remaining: {remaining} day(s)."
                return

        try:
            create_leave_request({
                "staff_email": email,
                "staff_name": fullname,
                "leave_type": leave_type,
                "start_date": start.strftime("%Y-%m-%d"),
                "end_date": end.strftime("%Y-%m-%d"),
                "days": days,
                "reason": reason,
                "submitted_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            ids.leave_message.text = "Leave request submitted successfully. Status: Pending."
            ids.leave_reason.text = ""
            self._refresh_leave_summary(email)
        except Exception as exc:
            logger.exception("Leave request submission failed")
            ids.leave_message.text = f"Could not submit request: {exc}"

    def export_leave_excel(self):
        """Export the logged-in staff member's leave history and balances to Excel."""
        try:
            if not self.current_user:
                return None
            email = str(self.current_user[20]) if len(self.current_user) > 20 else ""
            rows = get_leave_requests(email)
            headers = [
                "Leave Type", "Start Date", "End Date", "Days", "Reason",
                "Status", "Manager Comment", "Submitted At"
            ]
            data_rows = []
            for row in rows:
                # id, type, start, end, days, reason, status, manager_comment, submitted_at
                data_rows.append([
                    row[1] or "-", row[2] or "-", row[3] or "-", row[4] or 0,
                    row[5] or "-", row[6] or "Pending", row[7] or "-", row[8] or "-"
                ])

            # Keep a useful row even when the staff member has not requested leave yet.
            if not data_rows:
                data_rows.append(["No leave requests", "-", "-", 0, "-", "-", "-", "-"])

            reports_dir = self._get_download_dir()
            fullname = str(self.current_user[1]).replace(" ", "_") if len(self.current_user) > 1 else "staff"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(reports_dir, f"Leave_Report_{fullname}_{timestamp}.xlsx")
            self._build_excel(
                filepath, headers, data_rows,
                sheet_title="Leave Management", password=REPORT_EXPORT_PASSWORD
            )
            self._last_export_path = filepath
            self._queue_excel_auto_sync(filepath, "leave")
            if hasattr(self.leave_screen.ids, 'leave_message'):
                self.leave_screen.ids.leave_message.text = f"Leave report exported: {os.path.basename(filepath)}"
            logger.info("Leave report exported to %s", filepath)
            return filepath
        except Exception as exc:
            logger.exception("Error exporting leave report")
            if hasattr(self.leave_screen.ids, 'leave_message'):
                self.leave_screen.ids.leave_message.text = f"Could not export leave report: {exc}"
            return None

    # -----------------------------
    # Excel auto-sync connections and schedules
    # -----------------------------
    def _post_rohi_json(self, endpoint, payload, timeout=30):
        """POST structured ROHI data to an Apps Script/HTTP endpoint."""
        endpoint = str(endpoint or "").strip()
        if not endpoint:
            return False, "No automatic upload endpoint configured."
        try:
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            req = Request(endpoint, data=body, method="POST")
            req.add_header("Content-Type", "application/json; charset=utf-8")
            req.add_header("User-Agent", "ROHI-Attendance-App/2.0")
            with urlopen(req, timeout=timeout) as response:
                status = getattr(response, "status", 200)
                raw = response.read(8192).decode("utf-8", errors="replace")
            if status < 200 or status >= 300:
                return False, f"Server returned HTTP {status}."
            try:
                result = json.loads(raw) if raw else {}
            except Exception:
                result = {}
            if isinstance(result, dict) and result.get("ok") is False:
                return False, str(result.get("message") or "Server rejected the submission.")
            return True, str(result.get("message") if isinstance(result, dict) and result.get("message") else "Submitted successfully.")
        except HTTPError as exc:
            return False, f"Server submission failed (HTTP {exc.code})."
        except URLError:
            return False, "No internet connection. The submission will remain available for retry."
        except Exception as exc:
            logger.exception("ROHI JSON submission failed:")
            return False, f"Server submission failed: {exc}"

    def _submit_attendance_to_endpoint(self, check_in="", check_out="", checkin_gps="", checkout_gps=""):
        """Submit the current attendance row directly to the configured ROHI
        Attendance Apps Script endpoint. This is the Kobo-style data path:
        send row data, do not generate/upload an Excel file for every punch."""
        endpoint = str(self._excel_sync_state.get("attendance_endpoint") or "").strip()
        if not endpoint or not self.current_user:
            return False, "Attendance automatic endpoint is not configured."

        user = self.current_user
        payload = {
            "action": "attendance_submit",
            "staff_id": str(user[18] if len(user) > 18 else ""),
            "unique_id": str(user[29] if len(user) > 29 else ""),
            "name": str(user[1] if len(user) > 1 else ""),
            "sex": str(user[2] if len(user) > 2 else ""),
            "position": str(user[17] if len(user) > 17 else ""),
            "state_office": str(user[13] if len(user) > 13 else ""),
            "cluster": str(user[14] if len(user) > 14 else ""),
            "office_gps": str(user[26] if len(user) > 26 else self.static_gps),
            "date": datetime.now().strftime("%Y-%m-%d"),
            "check_in": check_in or "",
            "check_out": check_out or "",
            "checkin_gps": checkin_gps or "",
            "checkout_gps": checkout_gps or "",
        }
        return self._post_rohi_json(endpoint, payload, timeout=30)

    def _submit_staff_registration_to_endpoint(self, staff_data):
        """Immediately submit registration/profile data to the configured
        staff endpoint. This updates the Google Sheet directly; it is not an
        email action and does not require a Google account chooser."""
        endpoint = str(self._excel_sync_state.get("staff_endpoint") or "").strip()
        if not endpoint:
            return False, "Staff automatic endpoint is not configured."
        safe_data = dict(staff_data or {})
        # Never transmit the local password to the Google Sheet.
        safe_data.pop("password", None)
        safe_data["action"] = "staff_registration"
        safe_data["submitted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return self._post_rohi_json(endpoint, safe_data, timeout=30)

    def _submit_staff_registration_immediately(self, staff_data):
        """Run immediate staff registration sync in the background."""
        def worker():
            ok, message = self._submit_staff_registration_to_endpoint(staff_data)
            if ok:
                self._set_staff_registration_sync_status(
                    "Registration submitted to Google successfully.", True
                )
                return
            # Keep the existing workbook upload as a fallback when a file
            # endpoint is deliberately configured. Do not open Google or ask
            # the user to select an account.
            try:
                path = self._export_staff_template(for_sync=True)
                self._send_excel_now(path, "staff", self._set_staff_registration_sync_status)
            except Exception:
                logger.exception("Staff registration fallback export failed.")
                self._set_staff_registration_sync_status(message, False)
        threading.Thread(target=worker, daemon=True).start()

    def _send_excel_now(self, filepath, report_type, status_callback=None):
        """Send a generated workbook immediately.

        If an actual HTTP/Apps-Script upload endpoint is configured, the XLSX
        is uploaded in the background. If only a Google Sheets/Drive link is
        configured, that link is opened instead of incorrectly reporting
        "Not Connected". A share/folder URL cannot itself receive an XLSX POST,
        so the UI clearly tells the user that an upload endpoint is still needed
        for true unattended file transfer.
        """
        if not filepath or not os.path.exists(filepath):
            message = "Generate the Excel file first."
            if status_callback:
                Clock.schedule_once(lambda dt: status_callback(message, False), 0)
            return False

        endpoint = str(self._excel_sync_state.get(f"{report_type}_endpoint") or "").strip()

        # A Google Drive/Sheets view link is never treated as an upload API.
        # "Send to Google" must either upload through the configured Apps Script
        # endpoint or clearly report that the one-time endpoint configuration is
        # still missing. It must not open a browser or turn a link test into a
        # false upload success.
        if not endpoint:
            message = (
                f"{report_type.title()} Google upload is not configured yet. "
                "Paste the deployed ROHI Apps Script /exec URL into the "
                f"{report_type.title()} automatic upload endpoint in Server Connection."
            )
            if status_callback:
                Clock.schedule_once(lambda dt: status_callback(message, False), 0)
            return False

        def worker():
            ok, message = _http_upload_excel(filepath, endpoint, report_type)
            if ok:
                self._excel_sync_state[f"last_{report_type}_sync"] = datetime.now().isoformat(timespec="seconds")
                _save_excel_sync_config(self._excel_sync_state)
            if status_callback:
                Clock.schedule_once(lambda dt: status_callback(message, ok), 0)
        threading.Thread(target=worker, daemon=True).start()
        return True

    @mainthread
    def _set_report_send_status(self, message, ok=True):
        try:
            if hasattr(self.reports_screen.ids, "report_status_label"):
                self.reports_screen.ids.report_status_label.text = message
                self.reports_screen.ids.report_status_label.text_color = (0.13, 0.40, 0.16, 1) if ok else (0.8, 0.1, 0.1, 1)
        except Exception:
            pass

    def send_attendance_report_to_excel(self):
        """Send the already-generated attendance workbook to Google.

        This never generates the Excel report on its own — the user must
        press "Generate Report" first. If no matching report has been
        generated yet, show a clear message instead of creating one.
        """
        path = getattr(self, "_last_export_path", None)
        if not path or not os.path.exists(path) or "attendance" not in os.path.basename(path).lower():
            self._set_report_send_status("Generate the report first, then send it.", False)
            return
        self._send_excel_now(path, "attendance", self._set_report_send_status)

    @mainthread
    def _set_timesheet_send_status(self, message, ok=True):
        try:
            if hasattr(self.timesheet_screen.ids, "timesheet_status_label"):
                self.timesheet_screen.ids.timesheet_status_label.text = message
                self.timesheet_screen.ids.timesheet_status_label.text_color = (0.13, 0.40, 0.16, 1) if ok else (0.8, 0.1, 0.1, 1)
        except Exception:
            pass

    def send_timesheet_to_excel(self):
        """Send the already-exported timesheet workbook to Google.

        This never exports the Excel timesheet on its own — the user must
        press "Export Timesheet Excel" first. If no matching export exists
        yet, show a clear message instead of creating one.
        """
        path = getattr(self, "_last_export_path", None)
        if not path or not os.path.exists(path) or "timesheet" not in os.path.basename(path).lower():
            self._set_timesheet_send_status("Export the timesheet first, then send it.", False)
            return
        self._send_excel_now(path, "timesheet", self._set_timesheet_send_status)

    @mainthread
    def _set_leave_send_status(self, message, ok=True):
        try:
            if hasattr(self.leave_screen.ids, "leave_message"):
                self.leave_screen.ids.leave_message.text = message
                self.leave_screen.ids.leave_message.text_color = (0.13, 0.40, 0.16, 1) if ok else (0.8, 0.1, 0.1, 1)
        except Exception:
            pass

    def send_leave_to_excel(self):
        """Send the already-exported leave report workbook to Google.

        This never exports the Excel leave report on its own — the user
        must press "Export Leave Report Excel" first. If no matching export
        exists yet, show a clear message instead of creating one.
        """
        path = getattr(self, "_last_export_path", None)
        if not path or not os.path.exists(path) or "leave_report" not in os.path.basename(path).lower():
            self._set_leave_send_status("Export the leave report first, then send it.", False)
            return
        self._send_excel_now(path, "leave", self._set_leave_send_status)

    def _queue_excel_auto_sync(self, filepath, report_type):
        """Send an exported workbook to the configured server endpoint without
        blocking the UI. A plain Google Drive/Sheets share URL is only a link;
        actual upload requires an HTTP upload endpoint."""
        endpoint = str(self._excel_sync_state.get(f"{report_type}_endpoint") or "").strip()
        if not endpoint:
            return
        if not os.path.exists(filepath):
            return

        def worker():
            if not self._excel_sync_lock.acquire(blocking=False):
                logger.info("Excel sync already running; leaving current export on disk.")
                return
            try:
                ok, message = _http_upload_excel(filepath, endpoint, report_type)
                logger.info("Excel auto-sync [%s]: %s", report_type, message)
                if ok:
                    key = f"last_{report_type}_sync"
                    self._excel_sync_state[key] = datetime.now().isoformat(timespec="seconds")
                    _save_excel_sync_config(self._excel_sync_state)
            finally:
                self._excel_sync_lock.release()

        threading.Thread(target=worker, daemon=True).start()

    def _generate_and_push_staff_immediately(self, staff_data=None):
        """Immediately synchronize registration/profile data.

        The primary path is structured data -> Apps Script -> Google Sheet.
        The XLSX export remains only as a fallback when a file upload endpoint
        is configured. No Google account chooser is opened.
        """
        if staff_data:
            self._submit_staff_registration_immediately(staff_data)
            return None
        try:
            if self.current_user:
                data = self._staff_data_from_current_user()
                self._submit_staff_registration_immediately(data)
                return None
        except Exception:
            logger.exception("Immediate staff registration sync failed.")
        return None

    @mainthread
    def _set_staff_registration_sync_status(self, message, ok=True):
        logger.info("Staff registration sync status: %s", message)
        # Registration screen may have a status label in newer layouts. Keep
        # this optional so older layouts continue to work.
        try:
            ids = self.registration_screen.ids
            if "registration_sync_status" in ids:
                ids.registration_sync_status.text = message
                ids.registration_sync_status.text_color = (0.13, 0.40, 0.16, 1) if ok else (0.8, 0.1, 0.1, 1)
        except Exception:
            pass

    def _staff_data_from_current_user(self):
        """Build the non-secret registration payload from the current local record."""
        u = self.current_user or []
        def v(index):
            return str(u[index]) if len(u) > index and u[index] is not None else ""
        return {
            "fullname": v(1), "sex": v(2), "dob": v(3), "blood_group": v(4),
            "marital_status": v(5), "nationality": v(6), "state_origin": v(7),
            "lga": v(8), "address": v(9), "next_of_kin": v(10),
            "next_of_kin_phone": v(11), "employment_type": v(12),
            "state_office": v(13), "cluster": v(14), "department": v(15),
            "section": v(16), "position": v(17), "staff_number": v(18),
            "phone": v(19), "email": v(20), "facebook": v(21), "twitter": v(22),
            "instagram": v(23), "telegram": v(24), "linkedin": v(25),
            "gps_coordinate": v(26), "photo": v(27), "genotype": v(31),
            "reintegration_status": v(32), "unique_id": v(29),
        }

    def _generate_staff_excel_for_sync(self):
        try:
            path = self._export_staff_template(for_sync=True)
            self._queue_excel_auto_sync(path, "staff")
            self._last_staff_excel_sync_at = datetime.now().isoformat(timespec="seconds")
            self._excel_sync_state["last_staff_sync"] = self._last_staff_excel_sync_at
            _save_excel_sync_config(self._excel_sync_state)
            return path
        except Exception:
            logger.exception("24-hour staff Excel sync failed.")
            return None
        finally:
            self._excel_staff_schedule_running = False

    def _generate_timesheet_excel_for_sync(self):
        try:
            if not self.current_user:
                return None
            # _last_timesheet_rows is maintained by generate_timesheet() on the
            # UI thread. Exporting the workbook itself is safe in this worker.
            if not getattr(self, "_last_timesheet_rows", None):
                return None
            path = self._export_timesheet_template(for_sync=True)
            self._queue_excel_auto_sync(path, "timesheet")
            self._last_timesheet_excel_sync_at = datetime.now().isoformat(timespec="seconds")
            self._excel_sync_state["last_timesheet_sync"] = self._last_timesheet_excel_sync_at
            _save_excel_sync_config(self._excel_sync_state)
            return path
        except Exception:
            logger.exception("One-minute timesheet Excel sync failed.")
            return None
        finally:
            self._excel_timesheet_schedule_running = False

    def _generate_attendance_excel_for_sync(self):
        try:
            now = datetime.now()
            email = str(self.current_user[20]) if self.current_user and len(self.current_user) > 20 else ""
            rows = self._fetch_attendance_records(email, now.year, now.month, "All Days (Monthly)")
            today_rows = [r for r in rows if r[0] and r[0][:10] == now.strftime("%Y-%m-%d")]
            path = self._export_attendance_template(today_rows, "Daily")
            self._queue_excel_auto_sync(path, "attendance")
            self._last_attendance_excel_sync_date = now.strftime("%Y-%m-%d")
            self._excel_sync_state["last_attendance_sync"] = now.isoformat(timespec="seconds")
            _save_excel_sync_config(self._excel_sync_state)
            return path
        except Exception:
            logger.exception("5 PM attendance Excel sync failed.")
            return None
        finally:
            self._excel_attendance_schedule_running = False

    def _generate_leave_excel_for_sync(self):
        """Generate and queue the logged-in user's leave workbook every 3 minutes."""
        try:
            path = self.export_leave_excel()
            if path:
                self._last_leave_excel_sync_at = datetime.now().isoformat(timespec="seconds")
                self._excel_sync_state["last_leave_sync"] = self._last_leave_excel_sync_at
                _save_excel_sync_config(self._excel_sync_state)
        except Exception:
            logger.exception("Three-minute leave Excel auto-sync failed.")
        finally:
            self._excel_leave_schedule_running = False

    def push_all_reports_now(self):
        """Immediately generate and queue all configured report workbooks.

        The normal 3-minute auto-sync scheduler remains active. This button is
        simply an on-demand push and never disables the scheduler.
        """
        if not self.current_user:
            logger.warning("Immediate report push requested without an authenticated user.")
            return
        try:
            if hasattr(self.dashboard_screen.ids, "push_reports_btn"):
                self.dashboard_screen.ids.push_reports_btn.disabled = True
                self.dashboard_screen.ids.push_reports_btn.icon = "cloud-sync"
            if hasattr(self.dashboard_screen.ids, "push_reports_status"):
                self.dashboard_screen.ids.push_reports_status.text = "Pushing reports..."
        except Exception:
            pass

        def worker():
            results = []
            jobs = [
                ("attendance", self._generate_attendance_excel_for_sync),
                ("timesheet", self._generate_timesheet_excel_for_sync),
                ("leave", self._generate_leave_excel_for_sync),
                ("staff", self._generate_staff_excel_for_sync),
            ]
            for name, job in jobs:
                try:
                    path = job()
                    results.append(f"{name}: {'queued' if path else 'skipped'}")
                except Exception as exc:
                    logger.exception("Immediate %s report push failed", name)
                    results.append(f"{name}: failed")
            message = "All reports pushed/queued. Auto-sync remains every 3 minutes."
            Clock.schedule_once(lambda dt: self._finish_push_all_reports(message), 0)

        threading.Thread(target=worker, daemon=True).start()

    @mainthread
    def _finish_push_all_reports(self, message):
        try:
            if hasattr(self.dashboard_screen.ids, "push_reports_btn"):
                self.dashboard_screen.ids.push_reports_btn.disabled = False
                self.dashboard_screen.ids.push_reports_btn.icon = "cloud-upload"
            if hasattr(self.dashboard_screen.ids, "push_reports_status"):
                self.dashboard_screen.ids.push_reports_status.text = message
        except Exception:
            pass

    def _excel_sync_schedule_tick(self):
        """Run all Excel auto-sync jobs every 3 minutes while a user is logged in.

        The 30-second scheduler tick only checks whether each 180-second job is due;
        the actual Excel generation/upload runs in background threads so the UI stays responsive.
        """
        try:
            if not self.current_user:
                return
            now = datetime.now()

            def due(key, interval):
                last = str(self._excel_sync_state.get(key) or "")
                if not last:
                    return True
                try:
                    return (now - datetime.fromisoformat(last)).total_seconds() >= interval
                except Exception:
                    return True

            if due("last_staff_sync", STAFF_EXCEL_SYNC_INTERVAL_SECONDS) and not self._excel_staff_schedule_running:
                self._excel_staff_schedule_running = True
                threading.Thread(target=self._generate_staff_excel_for_sync, daemon=True).start()

            if due("last_timesheet_sync", TIMESHEET_EXCEL_SYNC_INTERVAL_SECONDS) and not self._excel_timesheet_schedule_running:
                self._excel_timesheet_schedule_running = True
                threading.Thread(target=self._generate_timesheet_excel_for_sync, daemon=True).start()

            if due("last_attendance_sync", ATTENDANCE_EXCEL_SYNC_INTERVAL_SECONDS) and not self._excel_attendance_schedule_running:
                self._excel_attendance_schedule_running = True
                threading.Thread(target=self._generate_attendance_excel_for_sync, daemon=True).start()

            if due("last_leave_sync", LEAVE_EXCEL_SYNC_INTERVAL_SECONDS) and not self._excel_leave_schedule_running:
                self._excel_leave_schedule_running = True
                threading.Thread(target=self._generate_leave_excel_for_sync, daemon=True).start()
        except Exception:
            logger.exception("Excel sync scheduler tick failed.")

    # -----------------------------
    # Timesheet (Excel export + signature upload)
    # -----------------------------
    def open_timesheet(self):
        self.root.current = "timesheet"
        if not self.current_user:
            return
        ids = self.timesheet_screen.ids
        fullname = str(self.current_user[1]) if len(self.current_user) > 1 and self.current_user[1] else "-"
        position = str(self.current_user[17]) if len(self.current_user) > 17 and self.current_user[17] else "-"
        if hasattr(ids, 'timesheet_name'):
            ids.timesheet_name.text = f"Name: {fullname}"
        if hasattr(ids, 'timesheet_position'):
            ids.timesheet_position.text = f"Position: {position}"
        now = datetime.now()
        ids.timesheet_month_spinner.text = now.strftime("%B")
        ids.timesheet_year_spinner.text = str(now.year)
        self.generate_timesheet()

    @staticmethod
    def _pay_period_days(year, month):
        """ROHI's payroll cycle runs 26th of the prior month through the 25th
        of the given month (matches the official timesheet template), rather
        than a plain calendar month."""
        end_date = datetime(year, month, 25)
        if month == 1:
            start_date = datetime(year - 1, 12, 26)
        else:
            start_date = datetime(year, month - 1, 26)
        days = []
        d = start_date
        while d <= end_date:
            days.append(d)
            d += timedelta(days=1)
        return days

    def _day_hours_worked(self, date_obj, record):
        """Returns the Hours Worked cell value for one pay-period day:
        WEEKEND / PUBLIC HOLIDAY labels, or fixed scheduled hours
        (Monday-Thursday: 9, Friday WFH: 7)."""
        date_key = date_obj.strftime("%Y-%m-%d")
        if date_key in ROHI_PUBLIC_HOLIDAYS:
            return "PUBLIC HOLIDAY"
        if date_obj.weekday() >= 5:  # Saturday=5, Sunday=6
            return "WEEKEND"

        # Friday is Work From Home: always counts as a full scheduled
        # 7-hour work day in the timesheet.
        if date_obj.weekday() == 4:
            return "4"

        # Monday-Thursday: always a scheduled 8-hour work day.
        return "8"

    def generate_timesheet(self):
        """Builds the daily timesheet grid (Pay Period Date | Day | Hours Worked)
        for the selected pay period (26th of prior month - 25th of selected
        month), matching the official ROHI timesheet template."""
        try:
            if not self.current_user:
                return
            ids = self.timesheet_screen.ids
            email_val = str(self.current_user[20]) if len(self.current_user) > 20 else ""
            month_name = ids.timesheet_month_spinner.text
            year = int(ids.timesheet_year_spinner.text)
            month = list(calendar.month_name).index(month_name)

            period_days = self._pay_period_days(year, month)
            start_d, end_d = period_days[0], period_days[-1]
            rows = self._fetch_attendance_records(email_val, start_d.year, start_d.month, "All Days (Monthly)")
            if start_d.month != end_d.month:
                rows += self._fetch_attendance_records(email_val, end_d.year, end_d.month, "All Days (Monthly)")

            by_date = {}
            for check_in, check_out, late, status, gps, checkout_gps in rows:
                if check_in:
                    by_date[check_in[:10]] = (check_in, check_out, late, status)

            container = ids.timesheet_rows_container
            container.clear_widgets()
            self._last_timesheet_rows = []
            total_hours = 0.0

            for date_obj in period_days:
                date_key = date_obj.strftime("%Y-%m-%d")
                day_name = date_obj.strftime("%A").upper()
                record = by_date.get(date_key)
                hours_cell = self._day_hours_worked(date_obj, record)
                try:
                    total_hours += float(hours_cell)
                except ValueError:
                    pass  # WEEKEND / PUBLIC HOLIDAY - not counted

                row = MDBoxLayout(orientation='horizontal', spacing=dp(4), size_hint_y=None, height=dp(26))
                row.add_widget(MDLabel(text=date_obj.strftime("%d/%m/%Y"), font_style="Caption"))
                row.add_widget(MDLabel(text=day_name, font_style="Caption"))
                row.add_widget(MDLabel(text=hours_cell, font_style="Caption"))
                container.add_widget(row)

                self._last_timesheet_rows.append([date_obj.strftime("%d/%m/%Y"), day_name, hours_cell])

            self._last_timesheet_total_hours = total_hours
            if hasattr(ids, 'timesheet_status_label'):
                ids.timesheet_status_label.text = (
                    f"Pay period {start_d.strftime('%d/%m/%Y')} - {end_d.strftime('%d/%m/%Y')} "
                    f"| Total: {total_hours:.0f} hrs"
                )
        except Exception:
            logger.exception("Error generating timesheet:")

    def upload_signature(self):
        """Select a staff signature from the phone gallery/file picker.

        This intentionally uses the same corrected Android picker as the
        registration photo control, including the fallback for builds where
        ``PythonActivity``/the activity wrapper has no ``bind`` method.
        """
        self._open_image_gallery('signature')

    def signature_captured(self, path):
        if path and os.path.exists(path):
            self.signature_path = path
            logger.info(f"Signature photo saved successfully to path: {path}")
            try:
                ids = self.timesheet_screen.ids
                if hasattr(ids, 'timesheet_signature_preview'):
                    ids.timesheet_signature_preview.source = path
                    ids.timesheet_signature_preview.reload()
            except Exception:
                logger.exception("Failed to refresh signature preview UI:")
        else:
            logger.warning("Signature capture returned no valid file path.")

    def export_timesheet_excel(self):
        try:
            filepath = self._export_timesheet_template()
            if not filepath:
                self.timesheet_screen.ids.timesheet_status_label.text = "Load a timesheet first."
                return None
            self._last_export_path = filepath
            self.timesheet_screen.ids.timesheet_status_label.text = (
                f"Exported: {os.path.basename(filepath)}"
            )
            self._queue_excel_auto_sync(filepath, "timesheet")
            logger.info("Timesheet exported from supplied template: %s", filepath)
            return filepath
        except Exception:
            logger.exception("Error exporting timesheet:")
            if hasattr(self.timesheet_screen.ids, "timesheet_status_label"):
                self.timesheet_screen.ids.timesheet_status_label.text = "Error exporting timesheet - see logs."
            return None



if __name__ == "__main__":
    try:
        ROHIAttendanceApp().run()
    except Exception:
        logger.exception("Application exited with unhandled top-level exception:")
        import traceback
        try:
            with open(os.path.join(APP_DIR, "logs", "startup_error.txt"), "a", encoding="utf-8") as f:
                f.write("\n===== STARTUP ERROR %s =====\n" % datetime.now())
                traceback.print_exc(file=f)
        except Exception: pass
        traceback.print_exc()
        raise
    finally:
        logger.info("ROHI Attendance App process stopped.")
